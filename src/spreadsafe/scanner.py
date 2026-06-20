from __future__ import annotations

from dataclasses import dataclass, field
from itertools import islice
from pathlib import Path
import csv
from typing import Any

from openpyxl import load_workbook

from spreadsafe.detectors import Detector, load_config
from spreadsafe.sanitizer import _looks_like_external_formula


@dataclass
class ColumnReport:
    name: str
    index: int
    inferred_type: str
    enum_values: list[str] = field(default_factory=list)
    null_count: int = 0
    sample_count: int = 0


@dataclass
class SheetReport:
    name: str
    hidden: bool
    rows: int
    columns: int
    headers: list[str]
    columns_report: list[ColumnReport]
    formulas: list[str] = field(default_factory=list)
    comments: list[str] = field(default_factory=list)
    data_validations: list[str] = field(default_factory=list)
    merged_cells: list[str] = field(default_factory=list)
    hyperlinks: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class WorkbookReport:
    path: str
    kind: str
    sheets: list[SheetReport]
    warnings: list[str] = field(default_factory=list)


def scan_directory(
    input_dir: Path,
    *,
    path_prefix: str | None = None,
    max_sample_rows_per_sheet: int = 500,
) -> list[WorkbookReport]:
    reports: list[WorkbookReport] = []
    for file_path in sorted(input_dir.rglob("*")):
        if file_path.is_dir():
            continue
        suffix = file_path.suffix.lower()
        report_path = _report_path(file_path, input_dir, path_prefix)
        if suffix == ".xlsx":
            reports.append(scan_xlsx(file_path, report_path, max_sample_rows_per_sheet))
        elif suffix == ".csv":
            reports.append(scan_csv(file_path, report_path, max_sample_rows_per_sheet))
        elif suffix in {".xls", ".xlsb", ".xlsm"}:
            reports.append(
                WorkbookReport(
                    path=report_path,
                    kind=suffix.removeprefix("."),
                    sheets=[],
                    warnings=[f"Unsupported workbook format: {suffix}"],
                )
            )
    return reports


def scan_xlsx(
    path: Path,
    report_path: str | None = None,
    max_sample_rows_per_sheet: int = 500,
) -> WorkbookReport:
    workbook = load_workbook(path, data_only=False)
    detector = Detector(load_config(None))
    warnings: list[str] = []
    if workbook.security is not None and workbook.security.lockStructure:
        warnings.append("Workbook structure is locked")
    sheets: list[SheetReport] = []
    for worksheet in workbook.worksheets:
        headers = [str(cell.value or f"Column {cell.column}") for cell in worksheet[1]]
        column_reports = _column_reports_from_rows(
            headers,
            worksheet.iter_rows(min_row=2, values_only=True),
            max_sample_rows_per_sheet,
        )
        formulas: list[str] = []
        comments: list[str] = []
        hyperlinks: list[str] = []
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    if detector.detect_text(cell.value):
                        formula = "[SENSITIVE_FORMULA]"
                    elif _looks_like_external_formula(cell.value):
                        formula = "[EXTERNAL_FORMULA]"
                    else:
                        formula = cell.value
                    formulas.append(f"{cell.coordinate}: {formula}")
                if cell.comment is not None:
                    comments.append(cell.coordinate)
                if cell.hyperlink is not None:
                    hyperlinks.append(cell.coordinate)
        validations = [str(validation.sqref) for validation in worksheet.data_validations.dataValidation]
        sheets.append(
            SheetReport(
                name=worksheet.title,
                hidden=worksheet.sheet_state != "visible",
                rows=worksheet.max_row,
                columns=worksheet.max_column,
                headers=headers,
                columns_report=column_reports,
                formulas=formulas,
                comments=comments,
                data_validations=validations,
                merged_cells=[str(cell_range) for cell_range in worksheet.merged_cells.ranges],
                hyperlinks=hyperlinks,
                warnings=["Hidden sheet"] if worksheet.sheet_state != "visible" else [],
            )
        )
    return WorkbookReport(path=report_path or path.name, kind="xlsx", sheets=sheets, warnings=warnings)


def scan_csv(
    path: Path,
    report_path: str | None = None,
    max_sample_rows_per_sheet: int = 500,
) -> WorkbookReport:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        try:
            headers = next(reader)
        except StopIteration:
            headers = []
            data_row_count = 0
            sample_rows: list[list[str]] = []
        else:
            data_row_count = 0
            sample_rows = []
            for row in reader:
                data_row_count += 1
                if len(sample_rows) < max_sample_rows_per_sheet:
                    sample_rows.append(row)
    column_reports = _column_reports_from_rows(headers, sample_rows, max_sample_rows_per_sheet)
    sheet = SheetReport(
        name=path.stem,
        hidden=False,
        rows=1 + data_row_count if headers else 0,
        columns=len(headers),
        headers=headers,
        columns_report=column_reports,
    )
    return WorkbookReport(path=report_path or path.name, kind="csv", sheets=[sheet])


def _report_path(file_path: Path, input_dir: Path, path_prefix: str | None) -> str:
    relative = file_path.relative_to(input_dir).as_posix()
    if path_prefix is None:
        return relative
    normalized_prefix = path_prefix.strip("/")
    return f"{normalized_prefix}/{relative}" if normalized_prefix else relative


def _column_reports_from_rows(
    headers: list[str],
    rows_iter: Any,
    max_sample_rows_per_sheet: int,
) -> list[ColumnReport]:
    rows = list(islice(rows_iter, max_sample_rows_per_sheet))
    reports: list[ColumnReport] = []
    for index, header in enumerate(headers):
        values = [row[index] if index < len(row) else None for row in rows]
        non_blank = [value for value in values if value not in (None, "")]
        enum_values = sorted({str(value) for value in non_blank if len(str(value)) <= 40})
        if len(enum_values) > 20:
            enum_values = []
        reports.append(
            ColumnReport(
                name=header,
                index=index + 1,
                inferred_type=_infer_type(non_blank),
                enum_values=enum_values,
                null_count=len(values) - len(non_blank),
                sample_count=len(non_blank),
            )
        )
    return reports


def _infer_type(values: list[Any]) -> str:
    if not values:
        return "empty"
    if all(isinstance(value, bool) for value in values):
        return "boolean"
    if all(isinstance(value, int | float) for value in values):
        return "number"
    if all(hasattr(value, "isoformat") for value in values):
        return "date"
    return "text"
