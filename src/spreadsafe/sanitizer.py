from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
import csv
import re
import shutil
from tempfile import TemporaryDirectory
from typing import Any
import zipfile

from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook
from openpyxl.worksheet.filters import AutoFilter

from spreadsafe.detectors import Config, Decision, Detection, Detector
from spreadsafe.mapping import PseudonymMapper


class Sanitizer:
    def __init__(self, config: Config, mapper: PseudonymMapper | None = None) -> None:
        self.config = config
        self.detector = Detector(config)
        self.mapper = mapper or PseudonymMapper()
        self.risks: list[str] = []

    def sanitize_directory(self, input_dir: Path, sanitized_dir: Path) -> None:
        sanitized_dir.mkdir(parents=True, exist_ok=True)
        used_destinations: set[Path] = set()
        for file_path in sorted(input_dir.rglob("*")):
            if file_path.is_symlink():
                self.risks.append("A symlinked input path was skipped")
                continue
            if file_path.is_dir():
                continue
            if file_path.name == "spreadsafe.yml":
                continue
            relative = self._sanitize_relative_path(file_path.relative_to(input_dir))
            destination = _unique_destination(sanitized_dir / relative, used_destinations)
            if destination.relative_to(sanitized_dir) != relative:
                self.risks.append("A sanitized destination name was adjusted to avoid collision")
            destination.parent.mkdir(parents=True, exist_ok=True)
            suffix = file_path.suffix.lower()
            if suffix == ".xlsx":
                self.sanitize_xlsx(file_path, destination)
            elif suffix == ".csv":
                self.sanitize_csv(file_path, destination)
            elif suffix in {".xls", ".xlsb", ".xlsm"}:
                self.risks.append(f"{relative}: unsupported format left out of sanitized output")
            else:
                self.risks.append(f"{relative}: unsupported file copied nowhere")

    def sanitize_xlsx(self, source: Path, destination: Path) -> None:
        with TemporaryDirectory(dir=destination.parent) as temporary:
            temporary_destination = Path(temporary) / destination.name
            shutil.copy2(source, temporary_destination)
            removed_unsupported_parts = _remove_unsupported_ooxml_parts(temporary_destination)
            workbook = load_workbook(temporary_destination, data_only=False, keep_links=False)
            _clear_document_properties(workbook)
            if _clear_custom_document_properties(workbook):
                self.risks.append(f"{source.name}: custom document properties removed")
            if workbook.defined_names:
                self.risks.append(f"{source.name}: workbook defined names removed")
                workbook.defined_names.clear()
            for worksheet in workbook.worksheets:
                if self.detector.detect_text(worksheet.title):
                    self.risks.append(f"{source.name}:{worksheet.title}: sheet title tokenized")
                    worksheet.title = self.mapper.token("SHEET", worksheet.title)
                if _clear_worksheet_headers_footers(worksheet):
                    self.risks.append(f"{source.name}:{worksheet.title}: print headers and footers removed")
                if _clear_auto_filter(worksheet):
                    self.risks.append(f"{source.name}:{worksheet.title}: auto filter removed")
                headers = [str(cell.value or f"Column {cell.column}") for cell in worksheet[1]]
                for cell in worksheet[1]:
                    if isinstance(cell.value, str):
                        cell.value = self._sanitize_value(
                            cell.value,
                            cell.value,
                            f"{source.name}:{worksheet.title}:{cell.coordinate}",
                        )
                if worksheet.sheet_state != "visible":
                    self.risks.append(f"{source.name}:{worksheet.title}: hidden sheet requires manual review")
                for row in worksheet.iter_rows(min_row=2):
                    for index, cell in enumerate(row):
                        if isinstance(cell, MergedCell):
                            continue
                        header = headers[index] if index < len(headers) else f"Column {index + 1}"
                        cell.value = self._sanitize_value(header, cell.value, f"{source.name}:{worksheet.title}:{cell.coordinate}")
                kept_validations = []
                for validation in worksheet.data_validations.dataValidation:
                    if _data_validation_has_external_reference(validation):
                        self.risks.append(f"{source.name}:{worksheet.title}: external data validation removed")
                    elif self.detector.detect_text(_data_validation_text(validation)):
                        self.risks.append(f"{source.name}:{worksheet.title}: sensitive data validation removed")
                    else:
                        kept_validations.append(validation)
                worksheet.data_validations.dataValidation = kept_validations
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.comment is not None:
                            self.risks.append(f"{source.name}:{worksheet.title}:{cell.coordinate}: comment removed")
                            cell.comment = None
                        if cell.hyperlink is not None:
                            self.risks.append(f"{source.name}:{worksheet.title}:{cell.coordinate}: hyperlink removed")
                            cell.hyperlink = None
            workbook.save(temporary_destination)
            _normalize_core_property_timestamps(temporary_destination)
            if _remove_external_link_parts(temporary_destination):
                self.risks.append(f"{source.name}: external workbook link metadata removed")
            if _remove_unsupported_ooxml_parts(temporary_destination) or removed_unsupported_parts:
                self.risks.append("Unsupported embedded workbook payloads were removed")
            temporary_destination.replace(destination)

    def sanitize_csv(self, source: Path, destination: Path) -> None:
        with TemporaryDirectory(dir=destination.parent) as temporary:
            temporary_destination = Path(temporary) / destination.name
            with (
                source.open(newline="", encoding="utf-8-sig") as input_handle,
                temporary_destination.open("w", newline="", encoding="utf-8") as output_handle,
            ):
                reader = csv.reader(input_handle)
                writer = csv.writer(output_handle)
                try:
                    headers = next(reader)
                except StopIteration:
                    pass
                else:
                    writer.writerow(
                        [
                            self._sanitize_header_value(header, source.name, index)
                            for index, header in enumerate(headers, start=1)
                        ]
                    )
                    for row_index, row in enumerate(reader, start=2):
                        if len(row) > len(headers):
                            raise ValueError(f"{source.name}: row {row_index} has extra columns")
                        if len(row) < len(headers):
                            row.extend([""] * (len(headers) - len(row)))
                        writer.writerow(
                            [
                                self._sanitize_csv_value(
                                    headers[index],
                                    value,
                                    source.name,
                                    row_index,
                                    index + 1,
                                )
                                for index, value in enumerate(row)
                            ]
                        )
            temporary_destination.replace(destination)

    def _sanitize_csv_value(
        self,
        header: str,
        value: Any,
        source_name: str,
        row_index: int,
        column_index: int,
    ) -> Any:
        if isinstance(value, str) and _looks_like_csv_formula(value):
            self.risks.append(
                f"{source_name}:row {row_index}:column {column_index}: CSV formula redacted"
            )
            return "[REDACTED_FORMULA]"
        coerced = _coerce_csv_value(header, value)
        sanitized = self._sanitize_value(
            header,
            coerced,
            f"{source_name}:row {row_index}:column {column_index}",
        )
        if isinstance(sanitized, str) and _looks_like_csv_formula(sanitized):
            self.risks.append(
                f"{source_name}:row {row_index}:column {column_index}: CSV formula redacted"
            )
            return "[REDACTED_FORMULA]"
        return sanitized

    def _sanitize_header_value(self, header: str, source_name: str, column_index: int) -> str:
        if _looks_like_csv_formula(header):
            self.risks.append(f"{source_name}:row 1:column {column_index}: CSV formula redacted")
            return "[REDACTED_FORMULA]"
        sanitized = str(
            self._sanitize_value(header, header, f"{source_name}:row 1:column {column_index}")
        )
        if _looks_like_csv_formula(sanitized):
            self.risks.append(f"{source_name}:row 1:column {column_index}: CSV formula redacted")
            return "[REDACTED_FORMULA]"
        return sanitized

    def _sanitize_relative_path(self, relative: Path) -> Path:
        parts = list(relative.parts)
        sanitized_parts: list[str] = []
        for directory in parts[:-1]:
            if self.detector.detect_path(directory):
                sanitized_parts.append(self.mapper.token("DIRECTORY", directory))
            else:
                sanitized_parts.append(directory)
        filename = parts[-1]
        suffix = Path(filename).suffix
        stem = Path(filename).stem
        if self.detector.detect_path(filename):
            filename = f"{self.mapper.token('FILE', stem)}{suffix}"
        sanitized_parts.append(filename)
        return Path(*sanitized_parts)

    def _sanitize_value(self, column_name: str, value: Any, location: str) -> Any:
        if value is None or value == "":
            return value
        if _column_matches(column_name, self.config.deny_columns):
            self.risks.append(f"{location}: redacted (config_deny_column)")
            return "[REDACTED_TEXT]"
        if isinstance(value, str) and value.startswith("="):
            decision = self.detector.classify_cell(column_name, value)
            if (
                decision.action != "preserve"
                or self.detector.detect_text(value)
                or _looks_like_external_formula(value)
            ):
                self.risks.append(f"{location}: formula redacted")
                return "[REDACTED_FORMULA]"
            return value
        decision = self.detector.classify_cell(column_name, value)
        if decision.action == "redact":
            self.risks.append(f"{location}: redacted ({', '.join(decision.reasons)})")
            return "[REDACTED_TEXT]"
        if decision.action == "tokenize":
            return self._tokenize(value, decision)
        if isinstance(value, datetime | date):
            return self.mapper.shift_date(value)
        if (
            isinstance(value, int | float | Decimal)
            and not isinstance(value, bool)
            and _looks_like_amount_column(column_name)
        ):
            return self.mapper.perturb_amount(value)
        return value

    def _tokenize(self, value: Any, decision: Decision) -> str:
        text = str(value)
        detections = _non_overlapping_detections(decision.detections)
        if len(detections) > 1 and decision.label not in {"COMPANY", "PERSON"}:
            replaced = text
            for detection in sorted(detections, key=lambda item: item.start, reverse=True):
                replacement = self.mapper.token(detection.label, detection.value)
                replaced = replaced[: detection.start] + replacement + replaced[detection.end :]
            return replaced
        if detections and decision.label not in {"COMPANY", "PERSON"}:
            label = detections[0].label
        else:
            label = decision.label or (detections[0].label if detections else "VALUE")
        return self.mapper.token(label, text)


def _looks_like_amount_column(column_name: str) -> bool:
    normalized = column_name.lower()
    return any(
        term in normalized
        for term in ("amount", "kwota", "price", "gross", "net", "total", "value", "wartosc")
    )


def _column_matches(column_name: str, candidates: list[str]) -> bool:
    normalized = column_name.strip().lower()
    return normalized in {candidate.strip().lower() for candidate in candidates}


def _looks_like_date_column(column_name: str) -> bool:
    normalized = column_name.lower()
    return any(term in normalized for term in ("date", "data", "due", "termin"))


def _coerce_csv_value(column_name: str, value: Any) -> Any:
    if not isinstance(value, str):
        return value
    stripped = value.strip()
    if not stripped:
        return value
    if _looks_like_date_column(column_name):
        parsed_date = _parse_iso_date(stripped)
        if parsed_date is not None:
            return parsed_date
    if _looks_like_amount_column(column_name):
        parsed_amount = _parse_decimal(stripped)
        if parsed_amount is not None:
            return parsed_amount
    return value


def _parse_iso_date(value: str) -> date | datetime | None:
    try:
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
            return date.fromisoformat(value)
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}(?::\d{2})?", value):
            return datetime.fromisoformat(value.replace(" ", "T"))
    except ValueError:
        return None
    return None


def _parse_decimal(value: str) -> Decimal | None:
    normalized = value.replace(" ", "").replace("\u00a0", "")
    if "," in normalized and "." in normalized:
        if normalized.rfind(",") > normalized.rfind("."):
            normalized = normalized.replace(".", "").replace(",", ".")
        else:
            normalized = normalized.replace(",", "")
    elif "," in normalized:
        normalized = normalized.replace(",", ".")
    try:
        return Decimal(normalized)
    except Exception:
        return None


def _looks_like_csv_formula(value: str) -> bool:
    stripped = value.lstrip()
    if not stripped:
        return False
    if stripped[0] in {"=", "+", "@"}:
        return True
    return stripped.startswith("-") and not _is_numeric_string(stripped)


def _looks_like_external_formula(value: str) -> bool:
    if not value.startswith("="):
        return False
    if re.search(
        r"(?:^|[^A-Z0-9_])(?:_xlfn\.)?(?:HYPERLINK|WEBSERVICE|FILTERXML|IMAGE)\s*\(",
        value,
        re.IGNORECASE,
    ):
        return True
    return _looks_like_external_reference(value)


def _looks_like_external_reference(value: str) -> bool:
    lowered = value.lower()
    if any(uri in lowered for uri in ("http://", "https://", "file://")):
        return True
    if re.search(r"\[[^\]]*(?:\.(?:xlsx|xlsm|xlsb|xls|csv)|[\\/]|:)[^\]]*\]", lowered):
        return True
    return False


def _data_validation_has_external_reference(validation: Any) -> bool:
    return any(
        isinstance(value, str) and _looks_like_external_reference(value)
        for value in _data_validation_values(validation)
    )


def _data_validation_text(validation: Any) -> str:
    return " ".join(
        str(value)
        for value in _data_validation_values(validation)
        if value is not None
    )


def _data_validation_values(validation: Any) -> tuple[Any, ...]:
    return (
        validation.formula1,
        validation.formula2,
        validation.promptTitle,
        validation.prompt,
        validation.errorTitle,
        validation.error,
    )


def _is_numeric_string(value: str) -> bool:
    parsed = _parse_decimal(value)
    return parsed is not None and parsed.is_finite()


def _non_overlapping_detections(detections: tuple[Detection, ...]) -> list[Detection]:
    ordered = sorted(
        detections,
        key=lambda item: (
            item.start,
            _detection_priority(item.label),
            -(item.end - item.start),
        ),
    )
    selected: list[Detection] = []
    for detection in ordered:
        if any(detection.start < existing.end and detection.end > existing.start for existing in selected):
            continue
        selected.append(detection)
    return selected


def _detection_priority(label: str) -> int:
    priorities = {
        "EMAIL": 0,
        "IBAN": 1,
        "PESEL": 2,
        "VAT_ID": 3,
        "REGON": 4,
        "NIP": 5,
        "INVOICE_ID": 6,
        "PHONE": 7,
        "COMPANY": 8,
    }
    return priorities.get(label, 99)


def _unique_destination(destination: Path, used_destinations: set[Path]) -> Path:
    if destination not in used_destinations and not destination.exists():
        used_destinations.add(destination)
        return destination
    counter = 2
    while True:
        candidate = destination.with_name(f"{destination.stem}_{counter}{destination.suffix}")
        if candidate not in used_destinations and not candidate.exists():
            used_destinations.add(candidate)
            return candidate
        counter += 1


def _clear_worksheet_headers_footers(worksheet: Any) -> bool:
    removed = False
    for container_name in (
        "oddHeader",
        "oddFooter",
        "evenHeader",
        "evenFooter",
        "firstHeader",
        "firstFooter",
    ):
        container = getattr(worksheet, container_name, None)
        if container is None:
            continue
        for section_name in ("left", "center", "right"):
            section = getattr(container, section_name, None)
            if section is not None and getattr(section, "text", None):
                section.text = None
                removed = True
    return removed


def _clear_auto_filter(worksheet: Any) -> bool:
    if not worksheet.auto_filter.ref and not worksheet.auto_filter.filterColumn:
        return False
    worksheet.auto_filter = AutoFilter()
    return True


def _clear_custom_document_properties(workbook: Any) -> bool:
    custom_properties = getattr(workbook, "custom_doc_props", None)
    properties = getattr(custom_properties, "props", None)
    if not properties:
        return False
    properties.clear()
    return True


def _clear_document_properties(workbook: Any) -> None:
    properties = workbook.properties
    properties.creator = "spreadsafe"
    properties.lastModifiedBy = "spreadsafe"
    properties.created = datetime(2000, 1, 1)
    properties.modified = datetime(2000, 1, 1)
    for field_name in (
        "title",
        "subject",
        "description",
        "keywords",
        "category",
        "contentStatus",
        "identifier",
        "language",
        "version",
        "revision",
    ):
        if hasattr(properties, field_name):
            setattr(properties, field_name, None)


def _normalize_core_property_timestamps(path: Path) -> None:
    entries = []
    changed = False
    with zipfile.ZipFile(path, "r") as source:
        for info in source.infolist():
            data = source.read(info.filename)
            if info.filename == "docProps/core.xml":
                for tag in (b"created", b"modified"):
                    cleaned = re.sub(
                        rb"(<dcterms:" + tag + rb"\b[^>]*>)[^<]*(</dcterms:" + tag + rb">)",
                        rb"\g<1>2000-01-01T00:00:00Z\2",
                        data,
                    )
                    changed = changed or cleaned != data
                    data = cleaned
            entries.append((info, data))
    if not changed:
        return
    with TemporaryDirectory() as temporary:
        temporary_path = Path(temporary) / path.name
        with zipfile.ZipFile(temporary_path, "w") as destination:
            for info, data in entries:
                destination.writestr(info, data)
        shutil.move(temporary_path, path)


def _remove_external_link_parts(path: Path) -> bool:
    removed = False
    with zipfile.ZipFile(path, "r") as source:
        entries: list[tuple[zipfile.ZipInfo, bytes]] = []
        for info in source.infolist():
            if info.filename.startswith("xl/externalLinks/"):
                removed = True
                continue
            data = source.read(info.filename)
            if info.filename == "xl/workbook.xml":
                cleaned = re.sub(
                    rb"<externalReferences\b[^>]*>.*?</externalReferences>",
                    b"",
                    data,
                    flags=re.DOTALL,
                )
                removed = removed or cleaned != data
                data = cleaned
            elif info.filename == "xl/_rels/workbook.xml.rels":
                cleaned = re.sub(
                    rb"<Relationship\b(?=[^>]*(?:externalLink|externalLinks/))[^>]*/>",
                    b"",
                    data,
                )
                removed = removed or cleaned != data
                data = cleaned
            entries.append((info, data))
    if not removed:
        return False
    with TemporaryDirectory() as temporary:
        temporary_path = Path(temporary) / path.name
        with zipfile.ZipFile(temporary_path, "w") as destination:
            for info, data in entries:
                destination.writestr(info, data)
        shutil.move(temporary_path, path)
    return True


def _remove_unsupported_ooxml_parts(path: Path) -> bool:
    removed = False
    with zipfile.ZipFile(path, "r") as source:
        unsupported_workbook_rel_ids = _unsupported_workbook_relationship_ids(source)
        entries: list[tuple[zipfile.ZipInfo, bytes]] = []
        for info in source.infolist():
            if _is_unsupported_ooxml_part(info.filename):
                removed = True
                continue
            data = source.read(info.filename)
            if info.filename.endswith(".rels"):
                cleaned = re.sub(
                    rb"<Relationship\b(?=[^>]*(?:/drawing|/image|/oleObject|/package|"
                    rb"/printerSettings|customXml|media/|embeddings/|charts/|chartsheets/|"
                    rb"/chart|/table|tables/|pivotCache|pivotTables?/))[^>]*/>",
                    b"",
                    data,
                )
                removed = removed or cleaned != data
                data = cleaned
            cleaned = _remove_unsupported_ooxml_references(
                info.filename,
                data,
                unsupported_workbook_rel_ids,
            )
            removed = removed or cleaned != data
            data = cleaned
            entries.append((info, data))
    if not removed:
        return False
    with TemporaryDirectory() as temporary:
        temporary_path = Path(temporary) / path.name
        with zipfile.ZipFile(temporary_path, "w") as destination:
            for info, data in entries:
                destination.writestr(info, data)
        shutil.move(temporary_path, path)
    return True


def _is_unsupported_ooxml_part(name: str) -> bool:
    return name.startswith(
        (
            "xl/chartsheets/",
            "xl/media/",
            "xl/drawings/",
            "xl/charts/",
            "xl/embeddings/",
            "xl/pivotCache/",
            "xl/pivotTables/",
            "xl/printerSettings/",
            "xl/tables/",
            "customXml/",
        )
    )


def _unsupported_workbook_relationship_ids(source: zipfile.ZipFile) -> set[bytes]:
    rels_name = "xl/_rels/workbook.xml.rels"
    if rels_name not in source.namelist():
        return set()
    rels = source.read(rels_name)
    return {
        match.group(1)
        for match in re.finditer(
            rb"<Relationship\b(?=[^>]*relationships/chartsheet)[^>]*\bId=\"([^\"]+)\"[^>]*/>",
            rels,
        )
    }


def _remove_unsupported_ooxml_references(
    name: str,
    data: bytes,
    unsupported_workbook_rel_ids: set[bytes],
) -> bytes:
    if name == "[Content_Types].xml":
        return re.sub(
            rb"<Override\b(?=[^>]*PartName=\"/(?:xl/(?:chartsheets|media|drawings|charts|embeddings|"
            rb"pivotCache|pivotTables|printerSettings|tables)/|customXml/))[^>]*/>",
            b"",
            data,
        )
    if name == "xl/workbook.xml":
        for rel_id in unsupported_workbook_rel_ids:
            data = re.sub(
                rb"<sheet\b(?=[^>]*\br:id=\"" + re.escape(rel_id) + rb"\")[^>]*/>",
                b"",
                data,
            )
        return re.sub(
            rb"<pivotCaches\b[^>]*>.*?</pivotCaches>",
            b"",
            data,
            flags=re.DOTALL,
        )
    if name.startswith("xl/worksheets/") and name.endswith(".xml"):
        cleaned = re.sub(
            rb"<(?:\w+:)?(?:drawing|legacyDrawing|legacyDrawingHF|picture)\b[^>]*/>",
            b"",
            data,
        )
        cleaned = re.sub(
            rb"<(?:\w+:)?oleObjects\b[^>]*>.*?</(?:\w+:)?oleObjects>",
            b"",
            cleaned,
            flags=re.DOTALL,
        )
        return re.sub(
            rb"<(?:\w+:)?tableParts\b[^>]*>.*?</(?:\w+:)?tableParts>",
            b"",
            cleaned,
            flags=re.DOTALL,
        )
    return data
