from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
import csv
import re
from typing import Any
import zipfile

from openpyxl import load_workbook

from spreadsafe.detectors import Config, Detection, Detector, load_config
from spreadsafe.sanitizer import (
    _data_validation_has_external_reference,
    _data_validation_text,
    _is_unsupported_ooxml_part,
    _looks_like_csv_formula,
    _looks_like_amount_column,
    _looks_like_date_column,
    _parse_decimal,
    _parse_iso_date,
)

ALLOWED_PACKAGE_ROOT_ENTRIES = {
    ".gitignore",
    ".spreadsafe-package",
    ".spreadsafe-reports",
    "reports",
    "sanitized",
}
SAFE_DOCUMENT_TIMESTAMP = datetime(2000, 1, 1)


@dataclass
class ValidationResult:
    passed: bool
    issues: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def validate_output(output_dir: Path, config: Config | None = None) -> ValidationResult:
    sanitized_dir = output_dir / "sanitized"
    issues: list[str] = []
    warnings: list[str] = []
    detector = Detector(config or load_config(None))
    output_location = _validation_location(output_dir.as_posix(), detector)
    sanitized_location = _validation_location(sanitized_dir.as_posix(), detector)
    if output_dir.is_symlink():
        issues.append(f"Package directory is a symlink: {output_location}")
    if output_dir.exists() and output_dir.is_dir():
        _validate_package_root(output_dir, detector, issues)
    if sanitized_dir.is_symlink():
        issues.append(f"Sanitized directory is a symlink: {sanitized_location}")
        return ValidationResult(False, issues, warnings)
    if not sanitized_dir.exists():
        issues.append(f"Missing sanitized directory: {sanitized_location}")
        return ValidationResult(False, issues, warnings)
    if not sanitized_dir.is_dir():
        issues.append(f"Sanitized path is not a directory: {sanitized_location}")
        return ValidationResult(False, issues, warnings)
    state_dir = output_dir / "state"
    if state_dir.exists():
        issues.append(
            "Package contains local re-identification state: "
            f"{_validation_location(state_dir.relative_to(output_dir).as_posix(), detector)}"
        )

    for file_path in sorted(sanitized_dir.rglob("*")):
        package_location = _validation_location(file_path.relative_to(output_dir).as_posix(), detector)
        if file_path.is_symlink():
            issues.append(f"Symlink is not allowed in sanitized package: {package_location}")
            continue
        if file_path.is_dir():
            continue
        _validate_path(file_path, sanitized_dir, detector, issues)
        if file_path.suffix.lower() == ".xlsx":
            _validate_xlsx(file_path, detector, issues, warnings)
        elif file_path.suffix.lower() == ".csv":
            _validate_csv_formulas(file_path, detector, issues)
            _validate_csv_values(file_path, detector, issues)
        else:
            issues.append(f"Unexpected sanitized file type: {package_location}")
    reports_dir = output_dir / "reports"
    reports_location = _validation_location(reports_dir.as_posix(), detector)
    if reports_dir.exists():
        if reports_dir.is_symlink():
            issues.append(f"Reports directory is a symlink: {reports_location}")
            return ValidationResult(False, issues, warnings)
        if reports_dir.is_dir():
            for file_path in sorted(reports_dir.rglob("*")):
                report_location = _validation_location(file_path.relative_to(output_dir).as_posix(), detector)
                if file_path.is_symlink():
                    issues.append(f"Symlink is not allowed in reports: {report_location}")
                    continue
                if file_path.is_dir():
                    issues.append(f"Unexpected reports directory: {report_location}")
                    continue
                if file_path.suffix.lower() in {".md", ".json", ".csv", ".txt"}:
                    _validate_text_file(file_path, detector, issues)
                else:
                    issues.append(f"Unexpected report file type: {report_location}")
    return ValidationResult(not issues, issues, warnings)


def _validate_package_root(output_dir: Path, detector: Detector, issues: list[str]) -> None:
    for child in sorted(output_dir.iterdir()):
        child_location = _validation_location(child.relative_to(output_dir).as_posix(), detector)
        if child.name not in ALLOWED_PACKAGE_ROOT_ENTRIES:
            issues.append(f"Unexpected package root entry: {child_location}")
            continue
        if child.is_symlink():
            issues.append(f"Symlink is not allowed in package root: {child_location}")
            continue
        if child.name == ".gitignore":
            if child.is_file():
                _validate_text_file(child, detector, issues)
            else:
                issues.append(f"Package .gitignore is not a file: {child_location}")
        elif child.name == ".spreadsafe-package":
            if not child.is_file():
                issues.append(f"Package marker is not a file: {child_location}")
            elif child.read_text(encoding="utf-8", errors="ignore") != "spreadsafe\n":
                issues.append(f"Package marker content is invalid: {child_location}")
        elif child.name == ".spreadsafe-reports":
            if not child.is_file():
                issues.append(f"Reports marker is not a file: {child_location}")
            elif child.read_text(encoding="utf-8", errors="ignore") != "spreadsafe-reports\n":
                issues.append(f"Reports marker content is invalid: {child_location}")
        elif child.name == "reports" and not child.is_dir():
            issues.append(f"Package reports path is not a directory: {child_location}")


def _validate_xlsx(
    file_path: Path,
    detector: Detector,
    issues: list[str],
    warnings: list[str],
) -> None:
    workbook_name = _validation_location(file_path.name, detector)
    for part_name in _external_link_part_names(file_path):
        issues.append(f"{workbook_name}: external workbook link metadata remains: {part_name}")
    unsupported_part_names = _unsupported_ooxml_part_names(file_path)
    for part_name in unsupported_part_names:
        issues.append(f"{workbook_name}: unsupported workbook payload remains: {part_name}")
    if unsupported_part_names:
        return
    workbook = load_workbook(file_path, data_only=False)
    for field_name, value in _document_property_values(workbook).items():
        for detection in detector.detect_text(value):
            if _is_safe_generated_value(detection.value):
                continue
            issues.append(
                f"{workbook_name}: document property {field_name} contains "
                f"{_validation_issue_label(detection)}"
            )
    for field_name in ("created", "modified"):
        timestamp_value = getattr(workbook.properties, field_name, None)
        if isinstance(timestamp_value, datetime) and timestamp_value != SAFE_DOCUMENT_TIMESTAMP:
            issues.append(f"{workbook_name}: document property {field_name} is not normalized")
    for property_name, value in _custom_document_property_values(workbook).items():
        property_location = _validation_location(property_name, detector)
        for detection in detector.detect_text(value):
            if _is_safe_generated_value(detection.value):
                continue
            issues.append(
                f"{workbook_name}: custom document property {property_location} contains "
                f"{_validation_issue_label(detection)}"
            )
    for name, defined_name in workbook.defined_names.items():
        name_location = _validation_location(name, detector)
        defined_text = " ".join(
            str(value)
            for value in (name, defined_name.attr_text, defined_name.comment, defined_name.description)
            if value is not None
        )
        for detection in detector.detect_text(defined_text):
            if _is_safe_generated_value(detection.value):
                continue
            issues.append(
                f"{workbook_name}: defined name {name_location} contains "
                f"{_validation_issue_label(detection)}"
            )
    for worksheet in workbook.worksheets:
        worksheet_name = _validation_location(worksheet.title, detector)
        headers = [str(cell.value or f"Column {cell.column}") for cell in worksheet[1]]
        for column_index, header in enumerate(headers, start=1):
            if _csv_header_is_sensitive(detector, header):
                issues.append(
                    f"{workbook_name}:{worksheet_name}: row 1 column {column_index} "
                    "contains sensitive header"
                )
        if not _is_safe_generated_value(worksheet.title):
            for detection in detector.detect_text(worksheet.title):
                issues.append(
                    f"{workbook_name}:{worksheet_name}: sheet title contains "
                    f"{_validation_issue_label(detection)}"
                )
        for location, value in _worksheet_header_footer_values(worksheet).items():
            for detection in detector.detect_text(value):
                if _is_safe_generated_value(detection.value):
                    continue
                issues.append(
                    f"{workbook_name}:{worksheet_name}: {location} contains "
                    f"{_validation_issue_label(detection)}"
                )
        if worksheet.sheet_state != "visible":
            warnings.append(f"{workbook_name}:{worksheet_name}: hidden sheet requires manual review")
        for validation in worksheet.data_validations.dataValidation:
            if _data_validation_has_external_reference(validation):
                issues.append(
                    f"{workbook_name}:{worksheet_name}: "
                    "data validation contains external workbook reference"
                )
            validation_text = _data_validation_text(validation)
            for detection in detector.detect_text(validation_text):
                if _is_safe_generated_value(detection.value):
                    continue
                issues.append(
                    f"{workbook_name}:{worksheet_name}: data validation contains "
                    f"{_validation_issue_label(detection)}"
                )
        for detection in detector.detect_text(_auto_filter_text(worksheet)):
            if _is_safe_generated_value(detection.value):
                continue
            issues.append(
                f"{workbook_name}:{worksheet_name}: auto filter contains "
                f"{_validation_issue_label(detection)}"
            )
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.comment is not None:
                    issues.append(f"{workbook_name}:{worksheet_name}:{cell.coordinate}: comment remains")
                if cell.hyperlink is not None:
                    issues.append(f"{workbook_name}:{worksheet_name}:{cell.coordinate}: hyperlink remains")
                value = cell.value
                if not isinstance(value, str):
                    header = headers[cell.column - 1] if cell.column - 1 < len(headers) else ""
                    if _value_violates_config_sensitive(detector, header, value):
                        issues.append(
                            f"{workbook_name}:{worksheet_name}:{cell.coordinate}: "
                            "configured sensitive value remains"
                        )
                        continue
                    if _is_safe_structured_xlsx_value(header, value):
                        continue
                    if value is not None:
                        for detection in detector.detect_text(str(value)):
                            if _is_safe_generated_value(detection.value):
                                continue
                            issues.append(
                                f"{workbook_name}:{worksheet_name}:{cell.coordinate}: "
                                f"{_validation_issue_label(detection)} remains"
                            )
                    continue
                if value.startswith("="):
                    issues.append(f"{workbook_name}:{worksheet_name}:{cell.coordinate}: formula remains")
                    continue
                header = headers[cell.column - 1] if cell.column - 1 < len(headers) else ""
                if _value_violates_config_sensitive(detector, header, value):
                    issues.append(
                        f"{workbook_name}:{worksheet_name}:{cell.coordinate}: "
                        "configured sensitive value remains"
                    )
                    continue
                for detection in detector.detect_text(value):
                    if _is_safe_generated_value(detection.value):
                        continue
                    issues.append(
                        f"{workbook_name}:{worksheet_name}:{cell.coordinate}: "
                        f"{_validation_issue_label(detection)} remains"
                    )


def _validate_text_file(file_path: Path, detector: Detector, issues: list[str]) -> None:
    file_name = _validation_location(file_path.name, detector)
    text = file_path.read_text(encoding="utf-8", errors="ignore")
    for detection in detector.detect_text(text):
        if _is_safe_generated_value(detection.value):
            continue
        issues.append(f"{file_name}: {_validation_issue_label(detection)} remains")


def _external_link_part_names(file_path: Path) -> list[str]:
    names: list[str] = []
    with zipfile.ZipFile(file_path, "r") as workbook_zip:
        for name in workbook_zip.namelist():
            if name.startswith("xl/externalLinks/"):
                names.append(name)
                continue
            if name == "xl/workbook.xml":
                content = workbook_zip.read(name)
                if b"<externalReferences" in content:
                    names.append(name)
            elif name == "xl/_rels/workbook.xml.rels":
                content = workbook_zip.read(name)
                if b"externalLink" in content or b"externalLinks/" in content:
                    names.append(name)
    return names


def _unsupported_ooxml_part_names(file_path: Path) -> list[str]:
    names: list[str] = []
    with zipfile.ZipFile(file_path, "r") as workbook_zip:
        for name in workbook_zip.namelist():
            if _is_unsupported_ooxml_part(name):
                names.append(name)
    return names


def _validate_path(
    file_path: Path,
    sanitized_dir: Path,
    detector: Detector,
    issues: list[str],
) -> None:
    relative = file_path.relative_to(sanitized_dir).as_posix()
    redacted_relative = _validation_location(relative, detector)
    for detection in detector.detect_path(relative):
        issues.append(f"{redacted_relative}: path contains {_validation_issue_label(detection)}")


def _validate_csv_formulas(file_path: Path, detector: Detector, issues: list[str]) -> None:
    file_name = _validation_location(file_path.name, detector)
    with file_path.open(newline="", encoding="utf-8-sig") as handle:
        for row_index, row in enumerate(csv.reader(handle), start=1):
            for column_index, value in enumerate(row, start=1):
                if _looks_like_csv_formula(value):
                    issues.append(
                        f"{file_name}: row {row_index} column {column_index} contains CSV formula"
                    )


def _validate_csv_values(file_path: Path, detector: Detector, issues: list[str]) -> None:
    file_name = _validation_location(file_path.name, detector)
    with file_path.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.reader(handle))
    headers = rows[0] if rows else []
    for row_index, row in enumerate(rows, start=1):
        if row_index > 1 and len(row) > len(headers):
            issues.append(f"{file_name}: row {row_index} contains extra columns")
        for column_index, value in enumerate(row, start=1):
            header = headers[column_index - 1] if row_index > 1 and column_index <= len(headers) else ""
            if row_index == 1 and _csv_header_is_sensitive(detector, value):
                issues.append(
                    f"{file_name}: row 1 column {column_index} contains sensitive header"
                )
            if row_index > 1 and _value_violates_config_sensitive(detector, header, value):
                issues.append(
                    f"{file_name}: row {row_index} column {column_index} "
                    "configured sensitive value remains"
                )
                continue
            if _is_safe_structured_csv_value(header, value):
                continue
            detections = detector.detect_text(value)
            if detections:
                for detection in detections:
                    if _is_safe_generated_value(detection.value):
                        continue
                    issues.append(
                        f"{file_name}: row {row_index} column {column_index} "
                        f"{_validation_issue_label(detection)} remains"
                    )
                continue


def _is_safe_structured_csv_value(header: str, value: str) -> bool:
    stripped = value.strip()
    if not stripped:
        return True
    if _looks_like_amount_column(header):
        parsed_amount = _parse_decimal(stripped)
        return (
            parsed_amount is not None
            and parsed_amount.is_finite()
            and any(separator in stripped for separator in (".", ","))
        )
    if _looks_like_date_column(header):
        return _parse_iso_date(stripped) is not None
    return False


def _is_safe_structured_xlsx_value(header: str, value: Any) -> bool:
    if isinstance(value, int) and not isinstance(value, bool):
        return False
    return isinstance(value, float) and _looks_like_amount_column(header) and not value.is_integer()


def _csv_header_is_sensitive(detector: Detector, value: str) -> bool:
    if _is_safe_generated_value(value) or value == "[REDACTED_TEXT]":
        return False
    decision = detector.classify_cell(value, value)
    return decision.action in {"redact", "tokenize"} and bool(
        {"column_heuristic", "config_sensitive_column"} & set(decision.reasons)
    )


def _value_violates_config_sensitive(detector: Detector, header: str, value: Any) -> bool:
    text = str(value)
    if not text.strip() or text in {"[REDACTED_TEXT]", "[REDACTED_FORMULA]"}:
        return False
    if _is_safe_generated_value(text):
        return False
    decision = detector.classify_cell(header, value)
    return decision.action in {"redact", "tokenize"} and "config_sensitive_column" in decision.reasons


def _is_safe_generated_value(value: str) -> bool:
    return bool(
        re.fullmatch(
            r"SPREADSAFE_(?:EMAIL|PHONE|IBAN|PESEL|NIP|REGON|VAT_ID|VALUE|INVOICE|"
            r"COMPANY|PERSON|SHEET)_\d{4}|"
            r"spreadsafe_(?:file|directory)_\d{4}",
            value,
        )
    )


def _validation_issue_label(detection: Detection) -> str:
    return detection.label


def _validation_location(value: str, detector: Detector) -> str:
    if "/" in value:
        return "/".join(_validation_location(part, detector) for part in value.split("/"))
    detections = [
        detection
        for detection in detector.detect_path(value)
        if not _is_safe_generated_value(detection.value)
    ]
    if not detections:
        return value
    chunks: list[str] = []
    cursor = 0
    for detection in sorted(detections, key=lambda item: (item.start, item.end)):
        if detection.start < cursor:
            continue
        chunks.append(value[cursor : detection.start])
        chunks.append(f"[{_validation_issue_label(detection)}]")
        cursor = detection.end
    chunks.append(value[cursor:])
    return "".join(chunks)


def _document_property_values(workbook: Any) -> dict[str, str]:
    properties = workbook.properties
    values: dict[str, str] = {}
    for field_name in (
        "creator",
        "lastModifiedBy",
        "title",
        "subject",
        "description",
        "keywords",
        "category",
        "contentStatus",
        "identifier",
        "language",
    ):
        value = getattr(properties, field_name, None)
        if isinstance(value, str) and value:
            values[field_name] = value
    return values


def _custom_document_property_values(workbook: Any) -> dict[str, str]:
    custom_properties = getattr(workbook, "custom_doc_props", None)
    values: dict[str, str] = {}
    for prop in getattr(custom_properties, "props", []):
        name = getattr(prop, "name", None)
        value = getattr(prop, "value", None)
        parts = [str(item) for item in (name, value) if item is not None]
        if name is not None and parts:
            values[str(name)] = " ".join(parts)
    return values


def _worksheet_header_footer_values(worksheet: Any) -> dict[str, str]:
    values: dict[str, str] = {}
    for container_name in (
        "oddHeader",
        "oddFooter",
        "evenHeader",
        "evenFooter",
        "firstHeader",
        "firstFooter",
    ):
        container = getattr(worksheet, container_name, None)
        if container is None:
            continue
        for section_name in ("left", "center", "right"):
            section = getattr(container, section_name, None)
            text = getattr(section, "text", None) if section is not None else None
            if isinstance(text, str) and text:
                values[f"{container_name}.{section_name}"] = text
    return values


def _auto_filter_text(worksheet: Any) -> str:
    auto_filter = worksheet.auto_filter
    parts: list[str] = []
    if isinstance(auto_filter.ref, str):
        parts.append(auto_filter.ref)
    parts.extend(str(filter_column) for filter_column in auto_filter.filterColumn)
    return " ".join(parts)
